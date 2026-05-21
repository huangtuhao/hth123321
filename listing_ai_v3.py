# poe: name=Listing-Designer

import openpyxl
import csv
import re
import json
import time
from io import BytesIO, StringIO
from fastapi_poe.types import SettingsResponse

poe.update_settings(SettingsResponse(
    introduction_message="""# 🚀 亚马逊Listing设计助手

欢迎使用！我将帮助您完成Listing的完整设计流程。

## 📋 第一步：信息收集

请上传以下信息：
1. **竞品数据Excel/CSV表格**（包含ASIN、标题、五点等信息）
2. **评论数据Excel/CSV表格**（竞品评论）
3. **参考产品链接**（完整URL，作为本产品的主要依据）
4. **产品图片**（可选，用于AI识别产品外观特征）
5. **产品描述**（可选，说明您的产品与参考竞品的不同之处）
6. **关键词数据Excel表格**（可选，用于标题五点优化，也可稍后上传）
7. **品牌名称**（可选，在消息中注明"品牌：XXX"即可自动识别）
8. **AI模型**（可选，默认Gemini-3.1-Pro，在消息中注明"AI：模型名"即可切换，例如"AI：Claude-Sonnet-4.6"）

💡 **提示**：
- 参考链接通常是最接近您产品的已上架产品
- 关键词数据如果现在不上传，选择生成「标题五点」时会再次提示
- 可以使用「参数：评论500，竞品50，关键词80」来调整数据读取量""",
))


def parse_json_from_response(response_text, required_fields=None):
    """从 LLM 响应中提取并解析 JSON，带容错处理。"""
    # 优先匹配 ```json 代码块
    match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
    if match:
        json_str = match.group(1).strip()
    else:
        # 回退：匹配任意 JSON 对象（贪婪，取最外层）
        match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if match:
            json_str = match.group(0).strip()
        else:
            raise Exception(f"无法在响应中找到 JSON:\n{response_text[:500]}")

    try:
        parsed = json.loads(json_str)
    except json.JSONDecodeError as e:
        # 容错：尝试修复常见 JSON 问题（尾部逗号等）
        cleaned = re.sub(r',\s*([}\]])', r'\1', json_str)
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            raise Exception(f"JSON 解析失败: {e}\n原始文本: {json_str[:500]}")

    if required_fields:
        missing = set(required_fields) - set(parsed.keys())
        if missing:
            raise Exception(f"JSON 缺少必需字段: {', '.join(missing)}")

    return parsed


class ListingDesigner:
    def __init__(self):
        # 基础阶段（1-3）
        self.base_stages = [
            "消费者洞察分析",
            "使用流程复现",
            "卖点提取"
        ]

        # 分支A：主副图（3个阶段）
        self.branch_a_stages = [
            "卖点选择与排序",
            "卖点表达方式设计",
            "主图+附图设计Brief生成"
        ]

        # 分支B：A+页面（2个阶段）
        self.branch_b_stages = [
            "A+页面模块规划",
            "A+页面设计Brief生成"
        ]

        # 分支C：标题五点（3个阶段）
        self.branch_c_stages = [
            "关键词策略分析",
            "标题生成",
            "五点描述生成"
        ]

    def run(self):
        chat_text = poe.default_chat.text
        user_input = poe.query.text.strip()

        # 检查是否是参数设置
        if user_input.startswith("参数：") or user_input.startswith("参数:"):
            self.handle_parameters(user_input)
            return

        # 检查是否已收集信息
        info_collected = "【信息已收集】" in chat_text
        base_completed = "【基础分析完成】" in chat_text

        # 情况1：信息收集阶段
        if not info_collected:
            self.handle_info_collection(user_input)
            return

        # 情况2：基础阶段（1-3）
        if not base_completed:
            self.handle_base_stage(user_input)
            return

        # 情况4：分支选择和执行
        self.handle_branch(user_input)

    def handle_parameters(self, user_input):
        """处理参数设置"""
        review_limit = 1000
        competitor_limit = 100
        keyword_limit = 100  # 新增：关键词限制

        if "评论" in user_input:
            match = re.search(r'评论\s*[:：]?\s*(\d+)', user_input)
            if match:
                review_limit = int(match.group(1))

        if "竞品" in user_input:
            match = re.search(r'竞品\s*[:：]?\s*(\d+)', user_input)
            if match:
                competitor_limit = int(match.group(1))

        if "关键词" in user_input:
            match = re.search(r'关键词\s*[:：]?\s*(\d+)', user_input)
            if match:
                keyword_limit = int(match.group(1))

        with poe.start_message() as msg:
            msg.write("## ⚙️ 参数设置成功\n\n")
            msg.write(f"📊 **竞品数据**：每个表格最多读取 {competitor_limit} 个\n")
            msg.write(f"💬 **评论数据**：每个表格最多读取 {review_limit} 条\n")
            msg.write(f"🔑 **关键词数据**：最多保留 {keyword_limit} 个（按搜索量排序）\n\n")
            msg.write(f"【参数设置：评论{review_limit}，竞品{competitor_limit}，关键词{keyword_limit}】\n\n")
            msg.write("💡 参数已设置！请上传数据文件开始分析。\n")

    def get_limits(self):
        """从历史中获取参数限制"""
        chat_text = poe.default_chat.text
        review_limit = 1000
        competitor_limit = 100
        keyword_limit = 100

        if "【参数设置：" in chat_text:
            match = re.search(r'【参数设置：评论(\d+)，竞品(\d+)(?:，关键词(\d+))?】', chat_text)
            if match:
                review_limit = int(match.group(1))
                competitor_limit = int(match.group(2))
                if match.group(3):
                    keyword_limit = int(match.group(3))

        return competitor_limit, review_limit, keyword_limit

    # ============ 摘要存储与读取 ============

    def store_summary(self, key, data):
        """将 JSON 摘要写入聊天历史，供后续阶段提取。

        存储格式: 【SUMMARY:key】{json}【/SUMMARY】
        标记由 Python 代码写入，不依赖 AI 输出，因此 100% 可靠。
        """
        json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        with poe.start_message() as msg:
            msg.write(f"【SUMMARY:{key}】{json_str}【/SUMMARY】\n")

    def get_summary(self, key):
        """从聊天历史中读取指定 key 的 JSON 摘要。

        返回最后一次存储的版本（支持重新生成覆盖）。
        找不到则返回 None。
        """
        chat_text = poe.default_chat.text
        pattern = f'【SUMMARY:{re.escape(key)}】(.*?)【/SUMMARY】'
        matches = re.findall(pattern, chat_text, re.DOTALL)
        if not matches:
            return None
        # 取最后一个匹配（最新的版本）
        try:
            return json.loads(matches[-1].strip())
        except json.JSONDecodeError:
            return None

    def get_all_summaries(self):
        """读取所有已存储的摘要，返回 {key: data} 字典。"""
        chat_text = poe.default_chat.text
        pattern = r'【SUMMARY:([\w.]+)】(.*?)【/SUMMARY】'
        matches = re.findall(pattern, chat_text, re.DOTALL)
        summaries = {}
        for key, json_str in matches:
            try:
                summaries[key] = json.loads(json_str.strip())
            except json.JSONDecodeError:
                continue
        return summaries

    def _format_stage_stats(self, stats):
        """格式化阶段统计信息，用于调试输出。"""
        lines = []

        elapsed = stats.get("elapsed_seconds", 0)
        total = stats.get("total_prompt_len", 0)
        output_len = stats.get("output_len", 0)

        # 粗略估算 token（中文约 1.5-2 字符/token，取 1.8）
        est_input_tokens = total / 1.8
        est_output_tokens = output_len / 1.8

        lines.append("📊 **本阶段统计：**")
        lines.append(f"- AI 耗时：**{elapsed:.1f} 秒**")
        lines.append(f"- 输入 prompt：{total:,} 字符（≈{est_input_tokens/1000:.1f}k tokens）")

        base_ctx_len = stats.get("base_context_len", 0)
        lines.append(f"  - 基础信息：{base_ctx_len:,} 字符")

        # 已知数据部分的总长度（用于计算指令部分）
        known_data_len = base_ctx_len

        # 注入的聚焦片段数据
        injected = stats.get("injected_data", [])
        if injected:
            injected_total = sum(d["len"] for d in injected)
            known_data_len += injected_total
            lines.append(f"  - 注入聚焦片段：{injected_total:,} 字符（{len(injected)}项）")
            for i, d in enumerate(injected):
                connector = "└" if i == len(injected) - 1 else "├"
                lines.append(f"    {connector} {d['name']}：{d['len']:,} 字符")
        else:
            lines.append("  - 注入聚焦片段：无（首阶段）")

        # 额外数据（关键词、VOC 等）—— extra_data 可能与 injected_data 重叠，
        # 这里只显示不在 injected_data 中的额外数据
        extra = stats.get("extra_data", [])
        injected_names = {d["name"] for d in injected}
        extra_unique = [e for e in extra if e["name"] not in injected_names]
        if extra_unique:
            for e in extra_unique:
                lines.append(f"  - {e['name']}：{e['len']:,} 字符")
                known_data_len += e["len"]

        # 阶段指令（总长 - 已知数据部分）
        feedback_len = stats.get("feedback_len", 0)
        instruction_len = total - known_data_len - feedback_len
        if instruction_len > 0:
            lines.append(f"  - 阶段指令：{instruction_len:,} 字符")
        if feedback_len > 0:
            lines.append(f"  - 用户反馈指令：{feedback_len:,} 字符")

        lines.append(f"- AI 输出：{output_len:,} 字符（≈{est_output_tokens/1000:.1f}k tokens）")

        return "\n".join(lines) + "\n"

    # 追加到每个 AI prompt 末尾，防止 AI 模仿生成系统标记
    _NO_MARKERS_INSTRUCTION = """

---

❌ **严禁在回答中输出任何【...】格式的系统流程标记。** 聊天记录中的【分支A开始】【基础阶段 1 完成】【分支B-阶段 2 完成】等标记是程序自动插入的流程控制符号，不属于你的输出内容。"""

    # 用于过滤 AI 输出中的系统标记
    _MARKER_RE = re.compile(
        r'【(?:'
        r'分支[A-Z]'
        r'|基础阶段'
        r'|基础分析'
        r'|信息已收集'
        r'|全部生成模式'
        r'|等待关键词'
        r'|SYS'
        r')[^】]*】'
    )
    _SYS_NO_MARKERS = (
        "系统规则：你的回答中绝对不能包含任何【...】格式的系统状态标记。"
        "聊天历史中出现的【分支A开始】【基础阶段 1 完成】【分支B-阶段 2 完成】等"
        "是程序自动生成的流程标记，你不能在回答中复制或模仿这些标记。只输出分析内容。"
    )

    def _call_ai(self, prompt):
        """调用AI并实时过滤输出中的系统标记。

        使用 poe.stream() + 缓冲过滤，防止 AI 生成假的状态标记污染聊天历史。
        返回过滤后的完整文本。
        """
        system_msg = poe.Message(text=self._SYS_NO_MARKERS, sender="system")
        marker_re = self._MARKER_RE
        buf_size = 40  # 缓冲区大小，覆盖最长标记

        output_parts = []
        with poe.start_message() as msg:
            buf = ""
            was_replace = False

            for p in poe.stream(
                self._get_model_name(),
                system_msg,
                poe.default_chat,
                prompt,
                adopt_current_bot_name=True
            ):
                if p.is_replace_response:
                    was_replace = True
                    buf = p.text
                    clean = marker_re.sub('', buf)
                    msg.overwrite(clean)
                    output_parts = [clean]
                else:
                    if was_replace:
                        was_replace = False
                        buf = ""
                    buf += p.text
                    if len(buf) > buf_size:
                        flush = buf[:-buf_size]
                        clean = marker_re.sub('', flush)
                        msg.write(clean)
                        output_parts.append(clean)
                        buf = buf[-buf_size:]

            # 刷出剩余缓冲
            if buf:
                clean = marker_re.sub('', buf)
                msg.write(clean)
                output_parts.append(clean)

        return ''.join(output_parts)

    def _get_keyword_data_for_prompt(self, top_n=None):
        """获取关键词数据的 markdown 表格，用于注入 prompt。

        Args:
            top_n: 只取前 N 个关键词（按搜索量排序）。None 表示全部。
        Returns:
            markdown 格式的关键词表格，或降级的旧式提取文本。
        """
        summary = self.get_summary("INFO_KEYWORDS")
        if summary and "keywords" in summary:
            keywords = summary["keywords"]
            stats = summary.get("stats", {})
            if top_n and len(keywords) > top_n:
                keywords = keywords[:top_n]
                header = f"**关键词数据（Top {top_n}，按搜索量排序）**\n\n"
            else:
                header = f"**关键词数据（共 {len(keywords)} 个，已筛选强相关+高相关）**\n\n"

            lines = [header]
            lines.append("| 序号 | 关键词 | 翻译 | 相关性 | 周搜索量 |")
            lines.append("|------|--------|------|--------|----------|")
            for i, kw in enumerate(keywords, 1):
                lines.append(f"| {i} | {kw.get('keyword', '')} | {kw.get('translation', '')} | {kw.get('relevance', '')} | {kw.get('volume', 0)} |")
            return "\n".join(lines)

        # 降级：旧式提取
        return self.extract_section("🔑 关键词数据")

    def _get_brand_name(self):
        """从聊天历史中获取用户的品牌名称。"""
        chat_text = poe.default_chat.text
        match = re.search(r'【品牌名：(.+?)】', chat_text)
        return match.group(1) if match else ""

    def _get_model_name(self):
        """从聊天历史中获取用户指定的AI模型名称，默认 Gemini-3.1-Pro。"""
        chat_text = poe.default_chat.text
        match = re.search(r'【AI模型：(.+?)】', chat_text)
        return match.group(1) if match else "Gemini-3.1-Pro"

    def _get_voc_for_prompt(self):
        """获取 VOC（买家原话）数据，用于注入五点生成 prompt。"""
        summary = self.get_summary("INFO_VOC")
        if not summary:
            return ""

        parts = ["**🗣️ 买家高频原话（Voice of Customer）— 请在五点中自然融入这些表达：**\n"]

        positive = summary.get("positive_phrases", [])
        if positive:
            parts.append(f"- 好评原话：{', '.join(positive[:15])}")

        negative = summary.get("negative_phrases", [])
        if negative:
            parts.append(f"- 差评原话（需在五点中正面回应）：{', '.join(negative[:10])}")

        scenario = summary.get("scenario_phrases", [])
        if scenario:
            parts.append(f"- 高频场景词：{', '.join(scenario[:10])}")

        emotional = summary.get("emotional_words", [])
        if emotional:
            parts.append(f"- 情感词汇：{', '.join(emotional[:10])}")

        return "\n".join(parts)

    def handle_info_collection(self, user_input):
        """处理信息收集阶段"""
        # 获取读取限制
        competitor_limit, review_limit, keyword_limit = self.get_limits()

        # 读取数据
        competitor_data = []
        review_data = []
        reference_links = []
        product_images = []
        product_description = ""
        keyword_data = []

        # 提取URL链接
        urls = re.findall(r'https?://[^\s]+', user_input)
        reference_links.extend(urls)

        # 提取产品描述
        remaining_text = re.sub(r'https?://[^\s]+', '', user_input).strip()
        if remaining_text:
            product_description = remaining_text

        # 自动提取品牌名称
        brand_name = ""
        if remaining_text:
            brand_patterns = [
                r'(?:品牌名[称]?|品牌)\s*[：:是为]\s*(.+?)(?:\n|$|[,，。.、;；])',
                r'(?:brand\s*name|brand)\s*[：:=]\s*(.+?)(?:\n|$|[,，。.、;；])',
            ]
            for pattern in brand_patterns:
                brand_match = re.search(pattern, remaining_text, re.IGNORECASE)
                if brand_match:
                    brand_name = brand_match.group(1).strip()
                    break

        # 自动提取AI模型名称
        ai_model = ""
        if remaining_text:
            ai_match = re.search(r'(?:AI|ai|Ai)\s*[：:]\s*([A-Za-z0-9._-]+)', remaining_text)
            if ai_match:
                ai_model = ai_match.group(1).strip()

        for attachment in poe.query.attachments:
            file_ext = attachment.name.lower()

            # 处理图片文件
            if file_ext.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                product_images.append({
                    'name': attachment.name,
                    'url': attachment.url
                })
                continue

            # 处理Excel文件
            if file_ext.endswith(('.xlsx', '.xls')):
                try:
                    contents = attachment.get_contents()
                    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))

                    if len(rows) > 1:
                        headers = rows[0]
                        headers_str = str(headers)
                        headers_upper = headers_str.upper()

                        # 关键词数据表（检测特征列）
                        if '关键词' in headers_str and ('搜索量' in headers_str or '流量' in headers_str):
                            for row in rows[1:]:
                                if row and any(row):
                                    kw_dict = dict(zip(headers, row))
                                    keyword_data.append(kw_dict)
                        # 竞品数据表
                        elif 'ASIN' in headers_upper and '评分数' in headers_str:
                            for row in rows[1:competitor_limit+1]:
                                if row and any(row):
                                    competitor_data.append(dict(zip(headers, row)))
                        # 评论数据表
                        elif 'ASIN' in headers_upper and ('评论' in headers_str or '内容' in headers_str or '星级' in headers_str):
                            for row in rows[1:review_limit+1]:
                                if row and any(row):
                                    review_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取Excel文件 {attachment.name} 时出错：{str(e)}\n")

            # 处理CSV文件
            elif file_ext.endswith('.csv'):
                try:
                    contents = attachment.get_contents()
                    try:
                        text = contents.decode('utf-8-sig')
                    except:
                        text = contents.decode('gbk')

                    csv_reader = csv.DictReader(StringIO(text))
                    rows = list(csv_reader)

                    if len(rows) > 0:
                        headers = list(rows[0].keys())
                        headers_upper = [h.upper() for h in headers]
                        has_asin = 'ASIN' in headers_upper
                        if has_asin and any(h in headers for h in ['Rating', 'Body', 'Title', 'ReviewID']):
                            for row in rows[:review_limit]:
                                review_dict = {
                                    'ASIN': row.get('ASIN', row.get('asin', '')),
                                    '标题': row.get('Title', row.get('标题', '')),
                                    '内容': row.get('Body', row.get('内容', '')),
                                    '星级': row.get('Rating', row.get('星级', '')),
                                    '评论时间': row.get('Date', row.get('评论时间', ''))
                                }
                                review_data.append(review_dict)
                        elif has_asin and any(h in headers for h in ['商品标题', '评分数', '月销量']):
                            for row in rows[:competitor_limit]:
                                competitor_data.append(row)
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取CSV文件 {attachment.name} 时出错：{str(e)}\n")

        # 如果有数据，进行处理
        if competitor_data or review_data or reference_links or product_images or product_description or keyword_data:
            with poe.start_message() as msg:
                msg.write("## ✅ 数据收集完成\n\n")
                msg.write(f"📊 **竞品数据**：{len(competitor_data)} 个竞品\n")
                msg.write(f"💬 **评论数据**：{len(review_data)} 条评论\n")
                msg.write(f"🔗 **参考链接**：{len(reference_links)} 个链接\n")
                msg.write(f"🖼️ **产品图片**：{len(product_images)} 张图片\n")
                msg.write(f"📝 **产品描述**：{'有' if product_description else '无'}\n")
                msg.write(f"🔑 **关键词数据**：{len(keyword_data)} 个关键词\n\n")
                msg.write("⏳ 正在对数据进行AI预处理...\n\n")

            # AI预处理评论数据
            if review_data:
                review_text = "\n".join([
                    f"[评分{review.get('星级', 'N/A')}] {review.get('标题', '')} {review.get('内容', '')}"
                    for review in review_data
                ])

                review_prompt = f"""基于以下{len(review_data)}条亚马逊评论，生成结构化数据摘要：

{review_text}

请输出以下内容：
1. **评分分布**：各星级数量统计
2. **高频好评要点**：按提及次数排序，列出前10个
3. **高频差评要点**：按提及次数排序，列出前10个
4. **使用场景总结**：用户在什么场景下使用
5. **典型评论**：挑选最有代表性的好评3条、中评2条、差评3条（原文引用）
6. **买家高频原话（Voice of Customer）**：提取评论中反复出现的英文原话/短语（至少15个），这些是买家自己的语言，后续将直接用于五点描述。格式如下：
   - 好评原话：[例如 "easy to clean", "fits perfectly", "keeps drinks cold all day"]
   - 差评原话：[例如 "lid is hard to open", "started leaking after a week"]
   - 高频场景词：[例如 "for my commute", "at the gym", "for hiking"]

要求：量化、直观、全面保留关键信息。买家原话必须是评论中真实出现的英文表达，不要翻译或改写。"""

                review_summary_response = poe.call(self._get_model_name(), review_prompt)

                with poe.start_message() as msg:
                    msg.write("### 📝 评论数据结构化摘要\n\n")
                    msg.write(review_summary_response.text)
                    msg.write("\n")

                # 新增：提取 VOC 原话并存储为 JSON
                voc_prompt = f"""从以下评论摘要中，提取买家高频原话（Voice of Customer）。

{review_summary_response.text[:4000]}

Output in ```json format:
```json
{{
    "positive_phrases": ["好评中的高频英文原话/短语，至少10个"],
    "negative_phrases": ["差评中的高频英文原话/短语，至少5个"],
    "scenario_phrases": ["使用场景相关的高频英文短语，至少5个"],
    "emotional_words": ["买家用来形容产品的高频形容词/情感词，至少5个"]
}}
```

要求：只提取评论中真实出现的英文表达，不要自己编造。"""

                try:
                    voc_response = poe.call("Gemini-3-Flash", voc_prompt)
                    voc_data = parse_json_from_response(voc_response.text, ["positive_phrases"])
                    self.store_summary("INFO_VOC", voc_data)
                except Exception:
                    pass  # VOC 提取失败不影响主流程

            # AI预处理竞品数据
            if competitor_data:
                comp_text = "\n".join([
                    f"{i+1}. 标题：{comp.get('商品标题', '')}\n   卖点：{comp.get('产品卖点', '')}"
                    for i, comp in enumerate(competitor_data)
                ])

                comp_prompt = f"""基于以下{len(competitor_data)}个竞品信息，生成结构化摘要：

{comp_text}

请输出：
1. **产品类型分析**：主要产品类型和特征
2. **高频卖点统计**：按出现频率排序，列出前15个常见卖点
3. **设计趋势**：材质、功能、风格等方面的趋势
4. **代表性产品**：挑选3个最有代表性的产品（包含标题和核心卖点）

要求：量化、直观、提炼关键信息。"""

                comp_summary_response = poe.call(self._get_model_name(), comp_prompt)

                with poe.start_message() as msg:
                    msg.write("### 🏆 竞品数据结构化摘要\n\n")
                    msg.write(comp_summary_response.text)
                    msg.write("\n")

            # 处理关键词数据（筛选+完整保留）
            if keyword_data:
                self.process_keyword_data(keyword_data, keyword_limit)

            # AI分析参考产品链接（使用web_search获取产品信息）
            if reference_links:
                self.analyze_reference_links(reference_links)

            # 显示其他收集的信息
            with poe.start_message() as msg:
                if product_description:
                    msg.write("### 📝 产品描述（与参考竞品的差异）\n\n")
                    msg.write(f"{product_description}\n\n")

                if product_images:
                    msg.write("### 🖼️ 产品图片（用于AI识别特征）\n\n")
                    for img in product_images:
                        msg.write(f"- 图片：{img['name']}\n")
                        msg.write(f"  - CDN链接：{img['url']}\n")
                    msg.write("\n")

                if brand_name:
                    msg.write(f"✅ 品牌名称已识别：**{brand_name}**\n\n")
                    msg.write(f"【品牌名：{brand_name}】\n\n")

                if ai_model:
                    msg.write(f"🤖 AI模型已切换：**{ai_model}**\n\n")
                    msg.write(f"【AI模型：{ai_model}】\n\n")

                msg.write("【信息已收集】\n\n")

            # 自动开始基础分析（消费者洞察→使用流程→卖点提取）
            self.execute_base_stage(0)

        else:
            with poe.start_message() as msg:
                msg.write("💡 请上传数据文件或粘贴链接。\n")

    def process_keyword_data(self, keyword_data, keyword_limit):
        """处理关键词数据：筛选+完整保留"""
        # 筛选强相关和高相关
        filtered_keywords = []
        for kw in keyword_data:
            relevance = kw.get('相关性档位', '')
            if relevance in ['强相关', '高相关']:
                filtered_keywords.append(kw)

        # 按周搜索量排序
        def get_search_volume(kw):
            vol = kw.get('周搜索量', 0)
            return int(vol) if vol else 0

        sorted_keywords = sorted(filtered_keywords, key=get_search_volume, reverse=True)

        # 限制数量
        final_keywords = sorted_keywords[:keyword_limit]

        # 生成摘要和完整列表
        with poe.start_message() as msg:
            msg.write("### 🔑 关键词数据（已筛选）\n\n")
            msg.write(f"**统计信息**\n")
            msg.write(f"- 原始词数：{len(keyword_data)} 个\n")
            msg.write(f"- 筛选后（强相关+高相关）：{len(filtered_keywords)} 个\n")
            msg.write(f"- 最终保留：{len(final_keywords)} 个\n\n")

            msg.write("**关键词完整列表（按搜索量排序）**\n\n")
            msg.write("| 序号 | 关键词 | 翻译 | 相关性 | 周搜索量 |\n")
            msg.write("|------|--------|------|--------|----------|\n")

            for i, kw in enumerate(final_keywords, 1):
                keyword = kw.get('关键词 (数据来源于西柚找词)', kw.get('关键词', ''))
                translation = kw.get('翻译', '')
                relevance = kw.get('相关性档位', '')
                volume = kw.get('周搜索量', 0)
                msg.write(f"| {i} | {keyword} | {translation} | {relevance} | {volume} |\n")

            msg.write("\n")

        # 新增：存储结构化 JSON，供后续精确切片
        keyword_json = []
        for kw in final_keywords:
            keyword_name = kw.get('关键词 (数据来源于西柚找词)', kw.get('关键词', ''))
            vol = kw.get('周搜索量', 0)
            keyword_json.append({
                "keyword": keyword_name,
                "translation": kw.get('翻译', ''),
                "relevance": kw.get('相关性档位', ''),
                "volume": int(vol) if vol else 0
            })
        self.store_summary("INFO_KEYWORDS", {
            "keywords": keyword_json,
            "stats": {
                "total_original": len(keyword_data),
                "total_filtered": len(filtered_keywords),
                "total_final": len(final_keywords)
            }
        })

    def analyze_reference_links(self, reference_links):
        """使用AI的web_search功能分析参考产品链接"""
        with poe.start_message() as msg:
            msg.write("### 🔗 参考产品分析\n\n")
            msg.write(f"正在通过网络搜索获取 {len(reference_links)} 个参考产品的详细信息...\n\n")

        links_text = "\n".join([f"- {link}" for link in reference_links])

        ref_prompt = f"""请访问以下亚马逊产品链接，获取并分析产品信息：

{links_text}

对于每个产品，请提取并整理以下信息：

## 参考产品1：[产品名称]
**链接**：[URL]

### 基本信息
- **完整标题**：[原文]
- **价格**：
- **评分**：
- **评论数**：

### 五点描述（Bullet Points）
1. [原文]
2. [原文]
3. [原文]
4. [原文]
5. [原文]

### 核心卖点分析
- **主打卖点**：[分析这个产品主打什么]
- **差异化特点**：[与同类产品的不同之处]
- **目标人群**：[面向什么用户]

### 主图分析
- **主图表达**：[主图展示了什么内容]
- **视觉风格**：[图片风格特点]

[如有多个链接，重复以上格式]

---

## 参考产品总结
- **共同特征**：[这些产品的共同点]
- **可借鉴之处**：[值得学习的地方]
- **可改进之处**：[可以做得更好的地方]

要求：
- 请务必访问链接获取真实信息，不要编造
- 尽可能提取完整的原文内容
- 分析要具体、有依据
"""

        # 使用AI模型的web_search功能（带重试）
        max_retries = 3
        for attempt in range(max_retries):
            try:
                ref_response = poe.call(
                    self._get_model_name(),
                    poe.Message(text=ref_prompt, parameters={"web_search": True})
                )
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 网络搜索出错，正在重试（{attempt + 2}/{max_retries}）...\n\n")
                    time.sleep(2)
                else:
                    raise

        with poe.start_message() as msg:
            msg.write(ref_response.text)
            msg.write("\n\n")
            msg.write("【参考产品分析完成】\n\n")

    def handle_base_stage(self, user_input):
        """处理基础阶段（1-3）"""
        chat_text = poe.default_chat.text
        current_base_stage = self.detect_base_stage()

        # 用户说继续，进入下一阶段
        if "继续" in user_input or "下一步" in user_input:
            next_stage = current_base_stage + 1
            if next_stage >= len(self.base_stages):
                # 基础阶段全部完成，显示分支菜单
                self.show_branch_menu()
            else:
                self.execute_base_stage(next_stage)
        elif "重新生成" in user_input:
            # 基于讨论内容重新生成当前阶段
            if current_base_stage >= 0:
                self.execute_stage_with_feedback(
                    "base", current_base_stage,
                    "请参考对话历史中关于当前阶段结果的讨论内容，重新生成。讨论中达成的共识和修改方向应体现在新的输出中。"
                )
            else:
                self.execute_base_stage(0)
        else:
            # 讨论模式：自由对话，不触发重新生成
            if current_base_stage >= 0:
                self.handle_discussion(user_input)
            else:
                self.execute_base_stage(0)

    def detect_base_stage(self):
        """检测当前基础阶段"""
        chat_text = poe.default_chat.text

        if "【信息已收集】" not in chat_text:
            return -1

        for i in range(len(self.base_stages) - 1, -1, -1):
            if f"【基础阶段 {i+1} 完成】" in chat_text:
                return i

        return -1

    def execute_base_stage(self, stage_num):
        """执行基础阶段"""
        with poe.start_message() as msg:
            msg.write(f"# ⚙️ 基础阶段 {stage_num + 1}：{self.base_stages[stage_num]}\n\n")
            msg.write("AI分析中，请稍候...\n\n")

        prompt, stats = self.build_base_prompt(stage_num)

        t0 = time.time()
        response_text = self._call_ai(prompt)
        stats["elapsed_seconds"] = time.time() - t0
        stats["output_len"] = len(response_text)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"【基础阶段 {stage_num + 1} 完成】{self.base_stages[stage_num]}\n\n")
            msg.write(self._format_stage_stats(stats))
            msg.write("\n")
            if stage_num < len(self.base_stages) - 1:
                msg.write("💡 直接输入讨论内容，或回复「**重新生成**」重做当前阶段，或回复「**继续**」进入下一阶段。\n")
            else:
                msg.write("💡 基础分析已完成！直接输入讨论内容，或回复「**重新生成**」重做当前阶段，或回复「**继续**」选择要生成的内容。\n")

    def show_branch_menu(self):
        """显示分支选择菜单"""
        chat_text = poe.default_chat.text

        with poe.start_message() as msg:
            msg.write("【基础分析完成】\n\n")
            msg.write("---\n\n")
            msg.write("## 🎯 请选择要生成的内容\n\n")
            msg.write("| 选项 | 内容 | 说明 |\n")
            msg.write("|------|------|------|\n")
            msg.write("| **1** | 主图+附图设计Brief | 7张图片的设计方案 |\n")
            msg.write("| **2** | A+页面设计Brief | A+页面模块设计方案 |\n")
            msg.write("| **3** | 标题+五点描述 | Listing文案优化 |\n")
            msg.write("| **4** | 全部生成 | 按顺序生成以上所有 |\n\n")

            # 检查已完成的分支
            completed = []
            if "【分支A完成】" in chat_text:
                completed.append("主图+附图")
            if "【分支B完成】" in chat_text:
                completed.append("A+页面")
            if "【分支C完成】" in chat_text:
                completed.append("标题五点")

            if completed:
                msg.write(f"✅ 已完成：{', '.join(completed)}\n\n")

            # 检查关键词数据
            if "🔑 关键词数据" not in chat_text:
                msg.write("⚠️ 提示：如需生成「标题+五点」，请先上传关键词数据表格。\n\n")

            msg.write("请输入数字（1-4）或名称选择：\n")

    def handle_branch(self, user_input):
        """处理分支选择和执行"""
        chat_text = poe.default_chat.text

        # 检查是否正在某个分支中
        current_branch, current_stage = self.detect_branch_stage()

        # 如果在分支中
        if current_branch:
            if "继续" in user_input or "下一步" in user_input:
                self.continue_branch(current_branch, current_stage)
            elif "重新生成" in user_input:
                self.execute_stage_with_feedback(
                    current_branch, current_stage,
                    "请参考对话历史中关于当前阶段结果的讨论内容，重新生成。讨论中达成的共识和修改方向应体现在新的输出中。"
                )
            else:
                # 讨论模式：自由对话，不触发重新生成
                self.handle_discussion(user_input)
            return

        # 检查是否需要上传关键词（选择了分支C但没有关键词数据）
        # 注意：如果已有关键词数据（🔑标记），则不再要求上传
        if "【等待关键词数据】" in chat_text and "🔑 关键词数据" not in chat_text:
            self.handle_keyword_upload(user_input)
            return

        # 解析用户选择
        choice = self.parse_branch_choice(user_input)

        if choice == "A":
            self.start_branch_a()
        elif choice == "B":
            self.start_branch_b()
        elif choice == "C":
            # 检查是否有关键词数据
            if "🔑 关键词数据" not in chat_text:
                self.request_keyword_upload()
            else:
                self.start_branch_c()
        elif choice == "ALL":
            # 标记全部生成模式
            with poe.start_message() as msg:
                msg.write("【全部生成模式】\n\n")
            self.start_branch_a()  # 先从A开始，完成后自动进入下一个
        else:
            # 无法识别，显示菜单
            self.show_branch_menu()

    def parse_branch_choice(self, user_input):
        """解析用户的分支选择"""
        user_input = user_input.lower()

        if "1" in user_input or "主图" in user_input or "附图" in user_input:
            return "A"
        elif "2" in user_input or "a+" in user_input or "A+" in user_input:
            return "B"
        elif "3" in user_input or "���题" in user_input or "五点" in user_input:
            return "C"
        elif "4" in user_input or "全部" in user_input or "都要" in user_input:
            return "ALL"

        return None

    def request_keyword_upload(self):
        """请求上传关键词数据"""
        with poe.start_message() as msg:
            msg.write("## 🔑 需要关键词数据\n\n")
            msg.write("生成「标题+五点描述」需要关键词数据表格。\n\n")
            msg.write("请上传包含以下列的Excel表格：\n")
            msg.write("- 关键词\n")
            msg.write("- 翻译（可选）\n")
            msg.write("- 相关性档位（强相关/高相关等）\n")
            msg.write("- 周搜索量\n\n")
            msg.write("【等待关键词数据】\n")

    def handle_keyword_upload(self, user_input):
        """处理关键词数据上传"""
        keyword_limit = self.get_limits()[2]
        keyword_data = []

        for attachment in poe.query.attachments:
            file_ext = attachment.name.lower()
            if file_ext.endswith(('.xlsx', '.xls')):
                try:
                    contents = attachment.get_contents()
                    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))

                    if len(rows) > 1:
                        headers = rows[0]
                        for row in rows[1:]:
                            if row and any(row):
                                keyword_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取文件出错：{str(e)}\n")

        if keyword_data:
            self.process_keyword_data(keyword_data, keyword_limit)
            # 检查分支C是否已经开始
            chat_text = poe.default_chat.text
            if "【分支C开始】" in chat_text and "【分支C完成】" not in chat_text:
                # 分支C已在进行中，继续当前阶段
                current_branch, current_stage = self.detect_branch_stage()
                if current_branch == "C":
                    self.continue_branch("C", current_stage)
                else:
                    # 某些异常情况，重新开始分支C
                    self.start_branch_c()
            else:
                # 分支C还没开始，开始它
                self.start_branch_c()
        else:
            # 检查是否实际上已经有关键词数据（用户可能只是发了不相关的消息）
            chat_text = poe.default_chat.text
            if "🔑 关键词数据" in chat_text:
                # 已有关键词数据，可能用户只是在聊天
                with poe.start_message() as msg:
                    msg.write("💡 关键词数据已存在。请选择要执行的操作，或回复「继续」进行下一步。\n")
            else:
                with poe.start_message() as msg:
                    msg.write("⚠️ 未检测到关键词数据，请上传Excel表格。\n")

    def detect_branch_stage(self):
        """检测当前所在的分支和阶段（检测最后一个活动的分支）"""
        chat_text = poe.default_chat.text

        # 找出所有活动的分支及其最后出现的位置
        active_branches = []

        # 检查分支C（优先检查，因为在全部生成模式下C是最后的）
        if "【分支C开始】" in chat_text and "【分支C完成】" not in chat_text:
            pos = chat_text.rfind("【分支C开始】")
            active_branches.append(("C", pos, self.branch_c_stages))

        # 检查分支B
        if "【分支B开始】" in chat_text and "【分支B完成】" not in chat_text:
            pos = chat_text.rfind("【分支B开始】")
            active_branches.append(("B", pos, self.branch_b_stages))

        # 检查分支A
        if "【分支A开始】" in chat_text and "【分支A完成】" not in chat_text:
            pos = chat_text.rfind("【分支A开始】")
            active_branches.append(("A", pos, self.branch_a_stages))

        if not active_branches:
            return None, -1

        # 选择位置最靠后的分支（最新开始的分支）
        active_branches.sort(key=lambda x: x[1], reverse=True)
        branch_name, _, stages = active_branches[0]

        # 找到该分支完成的最后一个阶段
        for i in range(len(stages) - 1, -1, -1):
            if f"【分支{branch_name}-阶段 {i+1} 完成】" in chat_text:
                return branch_name, i

        return branch_name, -1

    def start_branch_a(self):
        """开始分支A：主副图"""
        with poe.start_message() as msg:
            msg.write("## 🎨 开始生成：主图+附图设计Brief\n\n")
            msg.write("【分支A开始】\n\n")

        self.execute_branch_stage("A", 0)

    def start_branch_b(self):
        """开始分支B：A+页面"""
        with poe.start_message() as msg:
            msg.write("## 📄 开始生成：A+页面设计Brief\n\n")
            msg.write("【分支B开始】\n\n")

        self.execute_branch_stage("B", 0)

    def start_branch_c(self):
        """开始分支C：标题五点"""
        with poe.start_message() as msg:
            msg.write("## 📝 开始生成：标题+五点描述\n\n")
            msg.write("【分支C开始】\n\n")

        self.execute_branch_stage("C", 0)

    def continue_branch(self, branch, current_stage):
        """继续当前分支的下一阶段"""
        next_stage = current_stage + 1

        if branch == "A":
            stages = self.branch_a_stages
        elif branch == "B":
            stages = self.branch_b_stages
        else:
            stages = self.branch_c_stages

        if next_stage >= len(stages):
            self.complete_branch(branch)
        else:
            self.execute_branch_stage(branch, next_stage)

    def execute_branch_stage(self, branch, stage_num):
        """执行分支阶段"""
        if branch == "A":
            stages = self.branch_a_stages
        elif branch == "B":
            stages = self.branch_b_stages
        else:
            stages = self.branch_c_stages

        with poe.start_message() as msg:
            msg.write(f"# ⚙️ 分支{branch}-阶段 {stage_num + 1}：{stages[stage_num]}\n\n")
            msg.write("AI分析中，请稍候...\n\n")

        prompt, stats = self.build_branch_prompt(branch, stage_num)

        t0 = time.time()
        response_text = self._call_ai(prompt)
        stats["elapsed_seconds"] = time.time() - t0
        stats["output_len"] = len(response_text)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"【分支{branch}-阶段 {stage_num + 1} 完成】{stages[stage_num]}\n\n")
            msg.write(self._format_stage_stats(stats))
            msg.write("\n")

            if stage_num < len(stages) - 1:
                msg.write("💡 直接输入讨论内容，或回复「**重新生成**」重做当前阶段，或回复「**继续**」进入下一阶段。\n")
            else:
                msg.write("💡 直接输入讨论内容，或回复「**重新生成**」重做当前阶段，或回复「**继续**」完成此分支。\n")

    def complete_branch(self, branch):
        """完成分支"""
        chat_text = poe.default_chat.text
        is_all_mode = "【全部生成模式】" in chat_text

        # 先输出完成标记
        with poe.start_message() as msg:
            msg.write(f"【分支{branch}完成】\n\n")

            if branch == "A":
                msg.write("✅ 主图+附图设计Brief 已生成完成！\n\n")
            elif branch == "B":
                msg.write("✅ A+页面设计Brief 已生成完成！\n\n")
            else:
                msg.write("✅ 标题+五点描述 已生成完成！\n\n")

        # 检查是否是"全部生成"模式，继续下一个分支（在with块外调用）
        if is_all_mode:
            if branch == "A" and "【分支B完成】" not in chat_text:
                with poe.start_message() as msg:
                    msg.write("正在继续生成下一项...\n\n")
                self.start_branch_b()
                return
            elif branch == "B" and "【分支C完成】" not in chat_text:
                if "🔑 关键词数据" in chat_text:
                    with poe.start_message() as msg:
                        msg.write("正在继续生成下一项...\n\n")
                    self.start_branch_c()
                    return
                else:
                    with poe.start_message() as msg:
                        msg.write("⚠️ 缺少关键词数据，无法生成标题五点。\n\n")

        # 显示菜单
        with poe.start_message() as msg:
            msg.write("---\n\n")
            msg.write("您可以：\n")
            msg.write("- 输入数字选择生成其他内容\n")
            msg.write("- 对当前结果提出修改意见\n\n")

        self.show_branch_menu()

    def handle_discussion(self, user_input):
        """讨论模式：与AI自由对话讨论当前阶段结果，不触发重新生成。"""
        self._call_ai(user_input)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write("💡 继续讨论，或回复「**重新生成**」基于讨论重做当前阶段，或回复「**继续**」进入下一阶段。\n")

    def execute_stage_with_feedback(self, stage_type, stage_num, user_feedback):
        """根据用户反馈重新执行阶段"""
        if stage_type == "base":
            stage_name = self.base_stages[stage_num]
            prompt, stats = self.build_base_prompt(stage_num)
            marker = f"【基础阶段 {stage_num + 1} 完成】{stage_name}"
        else:
            if stage_type == "A":
                stage_name = self.branch_a_stages[stage_num]
            elif stage_type == "B":
                stage_name = self.branch_b_stages[stage_num]
            else:
                stage_name = self.branch_c_stages[stage_num]
            prompt, stats = self.build_branch_prompt(stage_type, stage_num)
            marker = f"【分支{stage_type}-阶段 {stage_num + 1} 完成】{stage_name}"

        with poe.start_message() as msg:
            msg.write(f"## 🔄 根据您的反馈重新生成\n\n")
            msg.write(f"**您的要求**：{user_feedback}\n\n")
            msg.write("AI正在重新生成，请稍候...\n\n")

        feedback_section = f"""

---

**用户特别要求：**
{user_feedback}

请严格遵守用户的要求重新生成内容。特别注意：
- 保留所有详细信息，不要精简
- 如果用户要求增加内容，务必添加
- 如果用户要求修改某部分，只修改该部分，其他部分保持原样"""

        enhanced_prompt = prompt + feedback_section
        stats["feedback_len"] = len(feedback_section)
        stats["total_prompt_len"] = len(enhanced_prompt)

        t0 = time.time()
        response_text = self._call_ai(enhanced_prompt)
        stats["elapsed_seconds"] = time.time() - t0
        stats["output_len"] = len(response_text)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"{marker}\n\n")
            msg.write(self._format_stage_stats(stats))
            msg.write("\n")
            msg.write("💡 直接输入讨论内容，或回复「**重新生成**」重做当前阶段，或回复「**继续**」进入下一阶段。\n")

    # ============ 基础上下文 ============

    def get_base_context(self):
        """获取所有基础上下文信息（不含关键词，关键词仅在分支C中按需注入）"""
        reference_analysis = self.extract_reference_analysis()
        product_description = self.extract_section("📝 产品描述（与参考竞品的差异）")
        product_images = self.extract_section("🖼️ 产品图片")
        review_summary = self.extract_section("📝 评论数据结构化摘要")
        competitor_summary = self.extract_section("🏆 竞品数据结构化摘要")

        brand = self._get_brand_name()
        context_parts = ["# 基础信息（所有阶段均可参考）\n"]

        if brand:
            context_parts.append(f"**⚠️ 用户的品牌名称：{brand}**（请在所有内容中使用此品牌名）\n\n")

        # 参考产品分析（最重要，放在最前面）
        if reference_analysis:
            if brand:
                context_parts.append(f"## 参考竞品分析（⚠️ 这是竞品数据，不是用户的产品。用户的品牌是「{brand}」，不要使用竞品品牌名）\n")
            else:
                context_parts.append("## 参考竞品分析（⚠️ 这是竞品数据，不是用户的产品）\n")
            context_parts.append(f"{reference_analysis}\n")

        if product_description:
            context_parts.append("## 产品描述（与参考竞品的差异）\n")
            context_parts.append(f"{product_description}\n")

        if product_images:
            context_parts.append("## 产品图片\n")
            context_parts.append(f"{product_images}\n")

        if review_summary:
            context_parts.append("## 评论数据摘要\n")
            context_parts.append(f"{review_summary}\n")

        if competitor_summary:
            context_parts.append("## 竞品数据摘要\n")
            context_parts.append(f"{competitor_summary}\n")

        # 关键词数据已移除：仅在分支C（标题五点）中按需注入，
        # 避免在不需要关键词的阶段（消费者洞察、主副图、A+页面等）浪费 Token

        context_parts.append("---\n\n")

        return "\n".join(context_parts)

    def extract_reference_analysis(self):
        """提取参考产品分析内容（从🔗标记到【参考产品分析完成】）"""
        chat_text = poe.default_chat.text

        start_marker = "### 🔗 参考产品分析"
        end_marker = "【参考产品分析完成】"

        if start_marker not in chat_text:
            return ""

        start_pos = chat_text.find(start_marker)
        end_pos = chat_text.find(end_marker)

        if start_pos >= 0 and end_pos > start_pos:
            content = chat_text[start_pos:end_pos]
            # 清理开头的标记
            content = content.replace(start_marker, "").strip()
            # 清理"正在通过网络搜索..."提示
            if "正在通过网络搜索" in content:
                lines = content.split("\n")
                content = "\n".join([l for l in lines if "正在通过网络搜索" not in l])
            return content.strip()[:8000]  # 参考产品分析可能较长，给更多空间

        return ""

    def extract_section(self, section_marker):
        """从对话历史中提取特定段落"""
        chat_text = poe.default_chat.text

        if section_marker not in chat_text:
            return ""

        parts = chat_text.split(section_marker)
        if len(parts) > 1:
            content = parts[1]
            end_markers = ["###", "【", "---", "##"]
            min_pos = len(content)
            for marker in end_markers:
                pos = content.find(marker, 50)
                if pos > 0 and pos < min_pos:
                    min_pos = pos
            return content[:min_pos].strip()[:3000]

        return ""

    def extract_stage_output(self, stage_type, stage_num):
        """提取指定阶段的输出"""
        chat_text = poe.default_chat.text

        if stage_type == "base":
            marker = f"【基础阶段 {stage_num + 1} 完成】"
        else:
            marker = f"【分支{stage_type}-阶段 {stage_num + 1} 完成】"

        if marker not in chat_text:
            return "[未找到阶段输出]"

        marker_pos = chat_text.rfind(marker)
        before_marker = chat_text[:marker_pos]

        # 找到阶段开始位置
        if stage_type == "base":
            if stage_num == 0:
                start_marker = "【信息已收集】"
            else:
                start_marker = f"【基础阶段 {stage_num} 完成】"
        else:
            if stage_num == 0:
                start_marker = f"【分支{stage_type}开始】"
            else:
                start_marker = f"【分支{stage_type}-阶段 {stage_num} 完成】"

        if start_marker in before_marker:
            start_pos = before_marker.rfind(start_marker)
        else:
            start_pos = 0

        content = chat_text[start_pos:marker_pos]

        # 清理内容
        lines = content.split('\n')
        clean_lines = [line for line in lines if not any(x in line for x in ['【', '💡', '请回复', 'AI分析中', '# ⚙️'])]

        result = '\n'.join(clean_lines).strip()
        return result[:4000] if len(result) > 100 else "[未找到有效阶段输出]"

    def get_previous_branch_outputs(self):
        """获取已完成分支的输出摘要（跨分支参考信息），每个分支截断到1500字符。"""
        chat_text = poe.default_chat.text
        sections = []

        for branch, label in [("A", "主图+附图"), ("B", "A+页面"), ("C", "标题五点")]:
            marker = f"【分支{branch}完成】"
            if marker not in chat_text:
                continue

            start_marker = f"【分支{branch}开始】"
            start_pos = chat_text.rfind(start_marker)
            end_pos = chat_text.rfind(marker)

            if start_pos >= 0 and end_pos > start_pos:
                content = chat_text[start_pos:end_pos]
                # 清理系统标记和UI元素
                lines = content.split('\n')
                clean_lines = [l for l in lines if not any(
                    x in l for x in ['【', '💡', '请回复', 'AI分析中', '# ⚙️', '📊 **本阶段统计']
                )]
                clean = '\n'.join(clean_lines).strip()
                if len(clean) > 1500:
                    clean = clean[:1500] + "\n...[已截断]"
                sections.append(f"### 已完成分支{branch}（{label}）摘要\n{clean}")

        if not sections:
            return ""

        return "## 📋 已完成的其他分支（跨分支参考）\n\n" + "\n\n".join(sections) + "\n\n"

    # ============ Prompt构建 ============

    def build_base_prompt(self, stage_num):
        """构建基础阶段的prompt。返回 (prompt, stats)。

        数据注入策略：只注入直接前序阶段的输出片段（4000字符截断）作为聚焦引导，
        完整上下文由 poe.default_chat 提供。
        """
        base_context = self.get_base_context()
        injected_data = []

        if stage_num == 0:
            prompt = f"""{base_context}

---

# 🎯 本阶段任务：消费者洞察分析

你是一位资深的亚马逊运营专家和消费者洞察分析师。请基于上述基础信息，进行深度的消费者洞察分析。

## 输出要求

请按照以下结构输出分析结果（使用中文）：

## 1. 消费者画像
- 主要用户群体特征
- 使用场景
- 购买动机

## 2. 购物决策因素
- 最重要的3-5个决策因素
- 每个因素的重要程度和原因

## 3. 核心痛点
- 用户最关心的痛点（按优先级排序）
- 每个痛点的具体表现

## 4. 期望与需求
- 用户明确提出的需求
- 隐含的期望

要求：
- 直观直白，避免理想化表达
- 基于数据事实，不要主观臆测
- 优先级明确，有理有据
"""

        elif stage_num == 1:
            consumer_insights = self.extract_stage_output("base", 0)
            injected_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})

            prompt = f"""{base_context}

**消费者洞察分析结果（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

---

# 🎯 本阶段任务：使用流程复现

你是一位产品体验设计专家。请聚焦于完整复现产品的使用流程。

## 输出要求

请按照时间顺序，详细描述产品使用的完整过程，包括：

## 使用前
- 用户的准备工作
- 需要了解的信息
- 开箱/安装步骤

## 使用中
- 具体使用步骤（分步骤详细说明）
- 每个步骤的操作要点
- 可能遇到的场景

## 使用后
- 收纳/维护
- 清洁保养
- 长期使用注意事项

要求：
- 力求完整，不遗漏任何步骤
- 代入消费者视角，真实还原使用场景
- 每个步骤都要具体、可操作
- 直白表达，不要过于理想化
"""

        elif stage_num == 2:
            usage_flow = self.extract_stage_output("base", 1)
            injected_data.append({"name": "使用流程（聚焦片段）", "len": len(usage_flow)})

            prompt = f"""{base_context}

**使用流程分析结果（聚焦片段，完整内容见对话历史）：**
{usage_flow}

---

# 🎯 本阶段任务：卖点提取

你是一位产品竞争分析专家。请聚焦于基于使用流程的每个步骤，提取产品卖点。

## 输出要求

请按照使用流程的每个步骤，分析：
1. 该步骤涉及的产品功能或特性
2. 与竞品对比，可能的优势是什么
3. 可以表达的卖点

输出格式：

## 使用前
### 步骤1：[步骤名称]
- **涉及功能**：
- **竞品对比**：
- **可表达卖点**：

[重复其他步骤]

## 使用中
[同上格式]

## 使用后
[同上格式]

## 卖点汇总（按重要性排序）
1. [最重要的卖点]
2. [次要卖点]
...

要求：
- 卖点要具体、可量化
- 对比要客观、有依据
- 避免空洞的形容词
- 直白表达
- 基于竞品分析、消费者洞察、使用流程来推导卖点
"""

        else:
            prompt = "[未定义阶段]"

        prompt += self._NO_MARKERS_INSTRUCTION

        stats = {
            "base_context_len": len(base_context),
            "injected_data": injected_data,
            "total_prompt_len": len(prompt),
            "extra_data": [],
        }
        return prompt, stats

    def build_branch_prompt(self, branch, stage_num):
        """构建分支阶段的prompt。返回 (prompt, stats)。

        数据注入策略：base_context + 直接需要的前序片段（4000字符截断）作为聚焦引导，
        stage 0 额外注入已完成分支摘要（跨分支参考），完整上下文由 poe.default_chat 提供。
        """
        base_context = self.get_base_context()
        selling_points = self.extract_stage_output("base", 2)  # 卖点提取
        reference_section = self.get_previous_branch_outputs() if stage_num == 0 else ""

        if branch == "A":
            prompt, extra_data = self.build_branch_a_prompt(stage_num, base_context, selling_points, reference_section)
        elif branch == "B":
            prompt, extra_data = self.build_branch_b_prompt(stage_num, base_context, selling_points, reference_section)
        else:
            prompt, extra_data = self.build_branch_c_prompt(stage_num, base_context, selling_points, reference_section)

        if reference_section:
            extra_data.append({"name": "跨分支参考", "len": len(reference_section)})

        stats = {
            "base_context_len": len(base_context),
            "injected_data": extra_data,
            "total_prompt_len": len(prompt),
            "extra_data": extra_data,
        }
        return prompt, stats

    def build_branch_a_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支A：主副图。返回 (prompt, extra_data)。"""
        extra_data = []
        if stage_num == 0:
            # A1：卖点选择与排序
            consumer_insights = self.extract_stage_output("base", 0)
            extra_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})
            extra_data.append({"name": "卖点提取（聚焦片段）", "len": len(selling_points)})

            prompt = f"""{base_context}
{reference_section}
**消费者洞察（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

**卖点提取（聚焦片段，完整内容见对话历史）：**
{selling_points}

---

# 🎯 本阶段任务：卖点选择与排序（主图+附图）

你是一位亚马逊Listing优化专家。请为亚马逊主图+6张附图（共7张）确定表达内容和顺序。

## 输出要求

基于基础分析结果，请输出：

## 图片规划

### 主图（第1张）
- **表达卖点**：[最核心的卖点]
- **选择理由**：[为什么这个卖点最重要]

### 附图1（第2张）
- **表达卖点**：
- **选择理由**：

### 附图2（第3张）
- **表达卖点**：
- **选择理由**：

### 附图3（第4张）
- **表达卖点**：
- **选择理由**：

### 附图4（第5张）
- **表达卖点**：
- **选择理由**：

### 附图5（第6张）
- **表达卖点**：
- **选择理由**：

### 附图6（第7张）
- **表达卖点**：
- **选择理由**：

## 排序逻辑说明
[解释为什么按这个顺序排列]

要求：
- 结合消费者最关心的点进行优先排序
- 考虑视觉呈现的逻辑性和连贯性
- 确保覆盖核心购买决策因素
"""

        elif stage_num == 1:
            # A2：卖点表达方式设计
            image_plan = self.extract_stage_output("A", 0)
            extra_data.append({"name": "卖点选择与排序（聚焦片段）", "len": len(image_plan)})

            prompt = f"""{base_context}

**卖点选择与排序结果（聚焦片段，完整内容见对话历史）：**
{image_plan}

---

# 🎯 本阶段任务：卖点表达方式设计

你是一位创意设计策略专家，精通各行业的视觉表达方式。请为每张图片设计专业的表达方式和AI绘图prompt。

## 输出要求

基于上述图片规划，对于每张图片，请：
1. 研究该卖点在同类产品中的常见表达方式
2. 思考其他行业/品牌如何专业地表达类似卖点
3. 提供具体的设计建议
4. 生成详细的AI绘图prompt

⚠️ **重要约束：每张图的卖点必须严格遵循前序「卖点选择与排序」阶段中的卖点分配，不得更改或替换。**

输出格式：

## 主图（第1张）

### 对应卖点（必须与图片规划一致）
[从前序阶段的图片规划中复制该图的卖点]

### 表达方式研究
- **同类产品常见表达**：
- **跨行业优秀案例**：
- **设计建议**：

### AI绘图Prompt
```
[详细的prompt]
```

[对其余6张附图重复以上格式，每张图的卖点必须与图片规划一一对应]
"""

        elif stage_num == 2:
            # A3：设计Brief生成
            consumer_insights = self.extract_stage_output("base", 0)
            image_plan = self.extract_stage_output("A", 0)
            design_approach = self.extract_stage_output("A", 1)
            extra_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})
            extra_data.append({"name": "卖点选择与排序（聚焦片段）", "len": len(image_plan)})
            extra_data.append({"name": "表达方式设计（聚焦片段）", "len": len(design_approach)})

            top_keywords = self._get_keyword_data_for_prompt(top_n=10)
            keyword_section = f"\n**Top 关键词（图片文案应尽量包含核心关键词）：**\n{top_keywords}\n" if top_keywords and "未找到" not in top_keywords else ""
            if keyword_section:
                extra_data.append({"name": "Top 关键词", "len": len(keyword_section)})

            prompt = f"""{base_context}

**消费者洞察（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

**图片规划（聚焦片段，完整内容见对话历史）：**
{image_plan}

**设计方案（聚焦片段，完整内容见对话历史）：**
{design_approach}
{keyword_section}

---

# 🎯 本阶段任务：主图+附图设计Brief生成

你是一位专业的设计项目经理。请生成一份完整的、适合外包设计公司使用的主图+附图设计Brief文档。

## 输出要求

请基于上述所有分析结果（特别是图片规划和设计方案），生成Brief文档。

⚠️ **重要约束：每张图的核心卖点必须严格遵循前序「卖点选择与排序」阶段中的卖点分配，不得更改、替换或重新排序。Brief 中每张图的核心卖点必须与图片规划一一对应。**

请生成一份结构清晰的Markdown格式设计Brief，包含：

# 亚马逊Listing图片设计Brief

## 1. 项目概述
- 产品类型
- 设计目标
- 交付要求（7张图片：1张主图+6张附图）

## 2. 消费者洞察总结

## 3. 图片设计需求详解

### 3.1 主图（第1张）
- **核心卖点**：
- **图片文案**：
  - **主标题**：[3-6个英文单词]
  - **副标题**：[可选]
- **构图设计**：
  - **画面布局**：
  - **拍摄角度**：
  - **产品状态**：
  - **背景设置**：
- **文字排版布局**：
  - **主标题位置**：
  - **���标题样式**：
- **视觉元素布局**：
- **色彩方案**：
- **AI绘图Prompt**：

[对其他6张图重复以上格式]

## 4. 设计规范
## 5. 验收标准清单

要求：
- 每张图片都必须有明确的英文文案
- **图片文案（Headline/Subheadline）应尽量自然融入核心关键词**，这既提升SEO也增强买家共鸣
- 构图设计必须详细具体
- 色彩方案必须明确
"""

        else:
            prompt = "[未定义阶段]"

        prompt += self._NO_MARKERS_INSTRUCTION
        return prompt, extra_data

    def build_branch_b_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支B：A+页面。返回 (prompt, extra_data)。"""
        extra_data = []
        brand = self._get_brand_name()
        brand_section = f"\n**⚠️ 品牌名称：{brand}**（A+页面中所有品牌相关内容必须使用此品牌名，绝对不要使用参考竞品的品牌名）\n" if brand else ""

        if stage_num == 0:
            # B1：A+页面模块规划
            consumer_insights = self.extract_stage_output("base", 0)
            usage_flow = self.extract_stage_output("base", 1)
            extra_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})
            extra_data.append({"name": "使用流程（聚焦片段）", "len": len(usage_flow)})
            extra_data.append({"name": "卖点提取（聚焦片段）", "len": len(selling_points)})

            prompt = f"""{base_context}
{reference_section}
{brand_section}

**消费者洞察（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

**使用流程（聚焦片段，完整内容见对话历史）：**
{usage_flow}

**卖点提取（聚焦片段，完整内容见对话历史）：**
{selling_points}

---

# 🎯 本阶段任务：A+页面模块规划

你是一位亚马逊A+页面设计专家。请为A+页面规划模块布局。

A+页面是品牌展示和深度说服的核心工具，需要通过多个模块组合来全面呈现产品价值。

⚠️ **图片规格要求**：所有 A+ 模块统一使用**高级A+完整图片模式**（Premium A+ Full-Width Image Module，每个模块为一张 1464×600px 的完整设计图，图片上集成文案和视觉元素，不使用亚马逊的默认文字+小图混排模板）。请基于此格式规划每个模块。

## 输出要求

请规划A+页面的模块设计（建议5个模块，可根据需要增减）：

## A+页面模块规划

### 模块1：[模块类型]
- **模块名称**：
- **核心目标**：
- **表达内容**：
- **图片构想**：[作为1464×600px的高级A+完整设计图，画面如何布局]

[重复其他模块]

## 模块排序逻辑
[解释为什么按这个顺序排列]
"""

        elif stage_num == 1:
            # B2：A+页面设计Brief
            module_plan = self.extract_stage_output("B", 0)
            extra_data.append({"name": "模块规划（聚焦片段）", "len": len(module_plan)})

            top_keywords = self._get_keyword_data_for_prompt(top_n=10)
            keyword_section = f"\n**Top 关键词（A+文案应自然融入，Amazon已对A+内容建立索引）：**\n{top_keywords}\n" if top_keywords and "未找到" not in top_keywords else ""
            if keyword_section:
                extra_data.append({"name": "Top 关键词", "len": len(keyword_section)})

            prompt = f"""{base_context}
{brand_section}

**A+页面模块规划（聚焦片段，完整内容见对话历史）：**
{module_plan}
{keyword_section}

---

# 🎯 本阶段任务：A+页面设计Brief生成

你是一位专业的A+页面设计项目经理。请生成详细的A+页面设计Brief文档。

## 输出要求

请基于上述模块规划，生成Brief文档。

⚠️ **图片规格要求**：所有模块统一使用**高级A+完整图片模式**（Premium A+ Full-Width Image Module，每个模块尺寸 1464×600px，文案和视觉元素直接设计在图片上）。

请生成：

# 亚马逊A+页面设计Brief

## 1. A+页面概述
- 设计目标
- 整体风格
- 模块总数
- **统一规格**：所有模块使用高级A+完整图片模式（1464×600px）

## 2. 模块设计需求详解

### 2.1 模块1：[模块名称]
- **模块格式**：高级A+完整图片模式（1464×600px 单张完整设计图）
- **核心目标**：
- **图片上的文案**：
  - **主标题**：[英文] — 位置和样式说明
  - **副标题**：[如有]
  - **正文/要点**：
- **画面布局**：[描述图片的整体构图]
- **视觉元素**：
- **色彩方案**：
- **AI绘图Prompt**：

[重复其他模块]

## 3. 设计规范
## 4. 验收标准
"""

        else:
            prompt = "[未定义阶段]"

        prompt += self._NO_MARKERS_INSTRUCTION
        return prompt, extra_data

    def build_branch_c_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支C：标题五点。返回 (prompt, extra_data)。"""
        extra_data = []
        if stage_num == 0:
            # C1：关键词策略分析 —— 使用完整关键词列表
            keyword_data = self._get_keyword_data_for_prompt()  # 全量
            extra_data.append({"name": "关键词数据（全量）", "len": len(keyword_data) if keyword_data else 0})
            extra_data.append({"name": "卖点提取（聚焦片段）", "len": len(selling_points)})

            prompt = f"""{base_context}
{reference_section}
**卖点提取（聚焦片段，完整内容见对话历史）：**
{selling_points}

{keyword_data}

---

# 🎯 本阶段任务：关键词策略分析

你是一位亚马逊SEO和关键词策略专家。请基于关键词数据和分析结果，制定关键词策略。

## 输出要求

基于上述卖点分析和关键词数据，请输出：

## 关键词策略分析

### 1. 核心主词（必须埋入标题）
| 关键词 | 周搜索量 | 选择理由 |
|--------|----------|----------|
| ... | ... | ... |

（建议3-5个核心主词）

### 2. 重要长尾词（标题或五点使用）
| 关键词 | 周搜索量 | 建议位置 |
|--------|----------|----------|
| ... | ... | 标题/五点第X条 |

（建议10-15个长尾词）

### 3. 五点专用词（按卖点分组）
| 卖点主题 | 相关关键词 | 建议融入方式 |
|----------|------------|--------------|
| 卖点1：XX | kw1, kw2 | ... |
| ... | ... | ... |

### 4. 关键词布局建议
- **标题**：[建议包含哪些词，大致顺序]
- **五点1**：[建议包含哪些词]
- **五点2**：[建议包含哪些词]
- **五点3**：[建议包含哪些词]
- **五点4**：[建议包含哪些词]
- **五点5**：[建议包含哪些词]
"""

        elif stage_num == 1:
            # C2：标题生成
            keyword_strategy = self.extract_stage_output("C", 0)
            consumer_insights = self.extract_stage_output("base", 0)
            extra_data.append({"name": "关键词策略（聚焦片段）", "len": len(keyword_strategy)})
            extra_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})
            extra_data.append({"name": "卖点提取（聚焦片段）", "len": len(selling_points)})

            prompt = f"""{base_context}

**关键词策略（聚焦片段，完整内容见对话历史）：**
{keyword_strategy}

**消费者洞察（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

**卖点提取（聚焦片段，完整内容见对话历史）：**
{selling_points}

---

# 🎯 本阶段任务：标题生成

你是一位亚马逊Listing标题优化专家。请基于关键词策略和卖点分析，生成标题方案。

## 输出要求

请输出：

## 标题生成

### 标题要求
- 字符限制：200字符以内
- 必须包含核心主词
- 突出最重要的1-2个卖点
- 符合亚马逊规范（无促销语、无特殊符号）

### 标题方案1（推荐）
**标题**：[完整标题]
**字符数**：[X字符]
**包含关键词**：[列出包含的核心词]
**突出卖点**：[列出突出的卖点]
**设计思路**：[解释为什么这样写]

### 标题方案2（备选）
[同上格式]

### 标题方案3（备选）
[同上格式]

### 方案对比
| 方案 | 优点 | 缺点 | 推荐度 |
|------|------|------|--------|
| 1 | ... | ... | ⭐⭐⭐⭐⭐ |
| 2 | ... | ... | ... |
| 3 | ... | ... | ... |
"""

        elif stage_num == 2:
            # C3：五点描述生成
            keyword_strategy = self.extract_stage_output("C", 0)
            title_output = self.extract_stage_output("C", 1)
            consumer_insights = self.extract_stage_output("base", 0)
            extra_data.append({"name": "关键词策略（聚焦片段）", "len": len(keyword_strategy)})
            extra_data.append({"name": "标题方案（聚焦片段）", "len": len(title_output)})
            extra_data.append({"name": "消费者洞察（聚焦片段）", "len": len(consumer_insights)})
            extra_data.append({"name": "卖点提取（聚焦片段）", "len": len(selling_points)})

            voc_data = self._get_voc_for_prompt()
            voc_section = f"\n{voc_data}\n" if voc_data else ""
            if voc_section:
                extra_data.append({"name": "VOC 数据", "len": len(voc_section)})

            prompt = f"""{base_context}

**关键词策略（聚焦片段，完整内容见对话历史）：**
{keyword_strategy}

**标题方案（聚焦片段，完整内容见对话历史）：**
{title_output}

**消费者洞察（聚焦片段，完整内容见对话历史）：**
{consumer_insights}

**卖点提取（聚焦片段，完整内容见对话历史）：**
{selling_points}
{voc_section}

---

# 🎯 本阶段任务：五点描述生成

你是一位亚马逊Listing五点描述优化专家。请基于关键词策略、标题方案和卖点分析，生成五点描述。

## 输出要求

请输出：

## 五点描述（Bullet Points）

### 五点要求
- 每条500字符以内
- 每条聚焦一个核心卖点
- 自然融入相关关键词
- 结构：卖点标题（大写） + 详细描述
- 符合亚马逊规范
- **重要：尽量使用买家原话（VOC）中的表达方式**，这些是真实买家的语言，比AI生成的表达更有说服力
- 如果差评原话中提到某个痛点，在相关五点中正面回应（例如差评说"lid is hard to open"，五点中可以写"EASY-OPEN magnetic lid..."）

### 五点1
**主题**：[卖点主题]
**包含关键词**：[列出融入的关键词]
**内容**：
```
[完整五点内容，含英文大写标题]
```
**字符数**：[X字符]

### 五点2
[同上格式]

### 五点3
[同上格式]

### 五点4
[同上格式]

### 五点5
[同上格式]

### 关键词覆盖检查
| 关键词 | 出现位置 |
|--------|----------|
| kw1 | 标题、五点1 |
| kw2 | 五点2 |
| ... | ... |

### Backend Search Terms（后台搜索词）
请推荐一组后台搜索词（总计不超过250字节），用于补充标题和五点中放不下的关键词。

**要求：**
- 不要重复标题和五点中已经出现的词
- 用空格分隔，不需要逗号
- 优先放入搜索量高但标题/五点中未使用的长尾词
- 可以包含常见拼写变体、同义词
- 不要包含品牌名、ASIN、"亚马逊"等禁用词

**推荐的 Backend Search Terms：**
```
[完整的后台搜索词，一行，空格分隔]
```
**字节数**：[X bytes / 250 bytes]
"""

        else:
            prompt = "[未定义阶段]"

        prompt += self._NO_MARKERS_INSTRUCTION
        return prompt, extra_data


if __name__ == "__main__":
    bot = ListingDesigner()
    bot.run()


