# poe: name=Listing-Designer

import openpyxl
import csv
import re
import json
from io import BytesIO, StringIO


def parse_json_from_response(response_text, required_fields=None):
    """从 LLM 响应中提取并解析 JSON，带容错处理。"""
    # 优先匹配 ```json 代码块
    match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
    if match:
        json_str = match.group(1).strip()
    else:
        # 回退：匹配任意 JSON 对象（贪婪，取最外层）
        match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if match:
            json_str = match.group(0).strip()
        else:
            raise Exception(f"无法在响应中找到 JSON:\n{response_text[:500]}")

    try:
        parsed = json.loads(json_str)
    except json.JSONDecodeError as e:
        # 容错：尝试修复常见 JSON 问题（尾部逗号等）
        cleaned = re.sub(r',\s*([}\]])', r'\1', json_str)
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            raise Exception(f"JSON 解析失败: {e}\n原始文本: {json_str[:500]}")

    if required_fields:
        missing = set(required_fields) - set(parsed.keys())
        if missing:
            raise Exception(f"JSON 缺少必需字段: {', '.join(missing)}")

    return parsed


class ListingDesigner:
    def __init__(self):
        # 基础阶段（1-3）
        self.base_stages = [
            "消费者洞察分析",
            "使用流程复现",
            "卖点提取"
        ]

        # 分支A：主副图（3个阶段）
        self.branch_a_stages = [
            "卖点选择与排序",
            "卖点表达方式设计",
            "主图+附图设计Brief生成"
        ]

        # 分支B：A+页面（2个阶段）
        self.branch_b_stages = [
            "A+页面模块规划",
            "A+页面设计Brief生成"
        ]

        # 分支C：标题五点（3个阶段）
        self.branch_c_stages = [
            "关键词策略分析",
            "标题生成",
            "五点描述生成"
        ]

    def run(self):
        chat_text = poe.default_chat.text
        user_input = poe.query.text.strip()

        # 检查是否是参数设置
        if user_input.startswith("参数：") or user_input.startswith("参数:"):
            self.handle_parameters(user_input)
            return

        # 检查是否已收集信息
        info_collected = "【信息已收集】" in chat_text
        base_completed = "【基础分析完成】" in chat_text

        # 情况1：信息收集阶段
        if not info_collected:
            self.handle_info_collection(user_input)
            return

        # 情况2：基础阶段（1-3）
        if not base_completed:
            self.handle_base_stage(user_input)
            return

        # 情况3：分支选择和执行
        self.handle_branch(user_input)

    def handle_parameters(self, user_input):
        """处理参数设置"""
        review_limit = 1000
        competitor_limit = 100
        keyword_limit = 100  # 新增：关键词限制

        if "评论" in user_input:
            match = re.search(r'评论\s*[:：]?\s*(\d+)', user_input)
            if match:
                review_limit = int(match.group(1))

        if "竞品" in user_input:
            match = re.search(r'竞品\s*[:：]?\s*(\d+)', user_input)
            if match:
                competitor_limit = int(match.group(1))

        if "关键词" in user_input:
            match = re.search(r'关键词\s*[:：]?\s*(\d+)', user_input)
            if match:
                keyword_limit = int(match.group(1))

        with poe.start_message() as msg:
            msg.write("## ⚙️ 参数设置成功\n\n")
            msg.write(f"📊 **竞品数据**：每个表格最多读取 {competitor_limit} 个\n")
            msg.write(f"💬 **评论数据**：每个表格最多读取 {review_limit} 条\n")
            msg.write(f"🔑 **关键词数据**：最多保留 {keyword_limit} 个（按搜索量排序）\n\n")
            msg.write(f"【参数设置：评论{review_limit}，竞品{competitor_limit}，关键词{keyword_limit}】\n\n")
            msg.write("💡 参数已设置！请上传数据文件开始分析。\n")

    def get_limits(self):
        """从历史中获取参数限制"""
        chat_text = poe.default_chat.text
        review_limit = 1000
        competitor_limit = 100
        keyword_limit = 100

        if "【参数设置：" in chat_text:
            match = re.search(r'【参数设置：评论(\d+)，竞品(\d+)(?:，关键词(\d+))?】', chat_text)
            if match:
                review_limit = int(match.group(1))
                competitor_limit = int(match.group(2))
                if match.group(3):
                    keyword_limit = int(match.group(3))

        return competitor_limit, review_limit, keyword_limit

    # ============ 摘要存储与读取 ============

    def store_summary(self, key, data):
        """将 JSON 摘要写入聊天历史，供后续阶段提取。

        存储格式: 【SUMMARY:key】{json}【/SUMMARY】
        标记由 Python 代码写入，不依赖 AI 输出，因此 100% 可靠。
        """
        json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        with poe.start_message() as msg:
            msg.write(f"📋 *阶段摘要已保存*\n\n")
            msg.write(f"【SUMMARY:{key}】{json_str}【/SUMMARY】\n")

    def get_summary(self, key):
        """从聊天历史中读取指定 key 的 JSON 摘要。

        返回最后一次存储的版本（支持重新生成覆盖）。
        找不到则返回 None。
        """
        chat_text = poe.default_chat.text
        pattern = f'【SUMMARY:{re.escape(key)}】(.*?)【/SUMMARY】'
        matches = re.findall(pattern, chat_text, re.DOTALL)
        if not matches:
            return None
        # 取最后一个匹配（最新的版本）
        try:
            return json.loads(matches[-1].strip())
        except json.JSONDecodeError:
            return None

    def get_all_summaries(self):
        """读取所有已存储的摘要，返回 {key: data} 字典。"""
        chat_text = poe.default_chat.text
        pattern = r'【SUMMARY:([\w.]+)】(.*?)【/SUMMARY】'
        matches = re.findall(pattern, chat_text, re.DOTALL)
        summaries = {}
        for key, json_str in matches:
            try:
                summaries[key] = json.loads(json_str.strip())
            except json.JSONDecodeError:
                continue
        return summaries

    def generate_and_store_summary(self, key, content, fields_description):
        """调用 AI 生成精简 JSON 摘要并存储。

        Args:
            key: 摘要标识符，如 'BASE_1', 'BRANCH_A_1'
            content: 需要被摘要的阶段输出全文
            fields_description: 告诉 AI 需要提取哪些字段的说明
        Returns:
            解析后的 JSON 字典，失败时返回包含 raw_excerpt 的降级字典
        """
        summary_prompt = f"""请将以下分析内容压缩为一份精简的 JSON 摘要。

**要求：**
- 只保留核心结论和关键数据点
- 去掉所有解释性文字、修饰语、过渡句
- 用中文
- 严格按照指定字段输出

**需要提取的字段：**
{fields_description}

Output in ```json format:
```json
{{...}}
```

---

以下是需要摘要的内容：

{content[:6000]}"""

        try:
            summary_response = poe.call("Gemini-3-Flash", summary_prompt)
            summary = parse_json_from_response(summary_response.text)
            self.store_summary(key, summary)
            return summary
        except Exception as e:
            # 降级处理：存储截断的原始文本
            fallback = {"raw_excerpt": content[:2000], "error": str(e)}
            self.store_summary(key, fallback)
            return fallback

    def _get_summary_fields_desc(self, stage_type, stage_num):
        """返回每个阶段需要的 JSON 摘要字段说明。"""
        fields_map = {
            ("base", 0): """- "consumer_personas": 主要用户群体及特征（字符串数组）
- "purchase_motivations": 核心购买动机（字符串数组）
- "decision_factors": 最重要的3-5个决策因素（字符串数组）
- "pain_points": 核心痛点，按优先级排序（字符串数组）
- "needs": 用户需求和期望（字符串数组）""",
            ("base", 1): """- "before_use": 使用前关键步骤（字符串数组）
- "during_use": 使用中关键步骤和操作要点（字符串数组）
- "after_use": 使用后维护/收纳要点（字符串数组）""",
            ("base", 2): """- "selling_points": 按重要性排序的卖点，每个元素为 {"name": "卖点名称", "description": "简述"}（对象数组）
- "competitive_advantages": 相对竞品的核心优势（字符串数组）""",
            ("A", 0): """- "image_plan": 7张图的卖点分配，每个元素为 {"image": "主图/附图N", "selling_point": "卖点", "reason": "理由"}（对象数组）
- "ordering_logic": 排序逻辑说明（字符串）""",
            ("A", 1): """- "design_approaches": 每张图的设计方案，每个元素为 {"image": "主图/附图N", "expression": "表达方式", "ai_prompt": "AI绘图prompt"}（对象数组）""",
            ("A", 2): """- "brief_summary": 设计Brief核心要点（字符串）
- "image_briefs": 每张图简要规格，每个元素为 {"image": "名称", "headline": "主标题", "layout": "布局"}（对象数组）""",
            ("B", 0): """- "modules": A+模块规划，每个元素为 {"name": "模块名", "goal": "目标", "content": "内容"}（对象数组）
- "ordering_logic": 排序逻辑（字符串）""",
            ("B", 1): """- "brief_summary": A+Brief核心要点（字符串）
- "module_briefs": 每个模块规格，每个元素为 {"name": "名称", "headline": "标题", "elements": "元素"}（对象数组）""",
            ("C", 0): """- "core_keywords": 核心主词（字符串数组）
- "longtail_keywords": 长尾词，每个元素为 {"keyword": "词", "location": "建议位置"}（对象数组）
- "keyword_layout": 标题和五点的关键词布局建议（字符串）""",
            ("C", 1): """- "recommended_title": 推荐标题全文（字符串）
- "title_keywords": 标题包含的关键词（字符串数组）
- "title_selling_points": 标题突出的卖点（字符串数组）""",
            ("C", 2): """- "bullet_points": 五点内容，每个元素为 {"theme": "主题", "content": "全文", "keywords": ["词1","词2"]}（对象数组）
- "keyword_coverage": 关键词覆盖摘要（字符串）""",
        }
        return fields_map.get((stage_type, stage_num), "- 请自行判断需要保留的核心信息字段")

    def _get_stage_summary_text(self, stage_type, stage_num):
        """获取阶段结论摘要文本，用于注入 prompt。

        优先使用 JSON 摘要（紧凑、可控），降级到旧式全文提取（兼容历史对话）。
        """
        if stage_type == "base":
            key = f"BASE_{stage_num + 1}"
        else:
            key = f"BRANCH_{stage_type}_{stage_num + 1}"

        summary = self.get_summary(key)
        if summary and "raw_excerpt" not in summary:
            return json.dumps(summary, ensure_ascii=False, indent=2)

        # 降级：使用旧式提取
        return self.extract_stage_output(stage_type, stage_num)

    def _get_keyword_data_for_prompt(self, top_n=None):
        """获取关键词数据的 markdown 表格，用于注入 prompt。

        Args:
            top_n: 只取前 N 个关键词（按搜索量排序）。None 表示全部。
        Returns:
            markdown 格式的关键词表格，或降级的旧式提取文本。
        """
        summary = self.get_summary("INFO_KEYWORDS")
        if summary and "keywords" in summary:
            keywords = summary["keywords"]
            stats = summary.get("stats", {})
            if top_n and len(keywords) > top_n:
                keywords = keywords[:top_n]
                header = f"**关键词数据（Top {top_n}，按搜索量排序）**\n\n"
            else:
                header = f"**关键词数据（共 {len(keywords)} 个，已筛选强相关+高相关）**\n\n"

            lines = [header]
            lines.append("| 序号 | 关键词 | 翻译 | 相关性 | 周搜索量 |")
            lines.append("|------|--------|------|--------|----------|")
            for i, kw in enumerate(keywords, 1):
                lines.append(f"| {i} | {kw.get('keyword', '')} | {kw.get('translation', '')} | {kw.get('relevance', '')} | {kw.get('volume', 0)} |")
            return "\n".join(lines)

        # 降级：旧式提取
        return self.extract_section("🔑 关键词数据")

    def handle_info_collection(self, user_input):
        """处理信息收集阶段"""
        chat_text = poe.default_chat.text
        is_first_run = "亚马逊Listing设计助手" not in chat_text

        if is_first_run:
            with poe.start_message() as msg:
                msg.write("# 🚀 亚马逊Listing设计助手\n\n")
                msg.write("欢迎使用！我将帮助您完成Listing的完整设计流程。\n\n")
                msg.write("## 📋 第一步：信息收集\n\n")
                msg.write("请上传以下信息：\n")
                msg.write("1. **竞品数据Excel/CSV表格**（包含ASIN、标题、五点等信息）\n")
                msg.write("2. **评论数据Excel/CSV表格**（竞品评论）\n")
                msg.write("3. **参考产品链接**（完整URL，作为本产品的主要依据）\n")
                msg.write("4. **产品图片**（可选，用于AI识别产品外观特征）\n")
                msg.write("5. **产品描述**（可选，说明您的产品与参考竞品的不同之处）\n")
                msg.write("6. **关键词数据Excel表格**（可选，用于标题五点优化，也可稍后上传）\n\n")
                msg.write("💡 **提示**：\n")
                msg.write("- 参考链接通常是最接近您产品的已上架产品\n")
                msg.write("- 关键词数据如果现在不上传，选择生成「标题五点」时会再次提示\n\n")

        # 获取读取限制
        competitor_limit, review_limit, keyword_limit = self.get_limits()

        # 读取数据
        competitor_data = []
        review_data = []
        reference_links = []
        product_images = []
        product_description = ""
        keyword_data = []

        # 提取URL链接
        urls = re.findall(r'https?://[^\s]+', user_input)
        reference_links.extend(urls)

        # 提取产品描述
        remaining_text = re.sub(r'https?://[^\s]+', '', user_input).strip()
        if remaining_text:
            product_description = remaining_text

        for attachment in poe.query.attachments:
            file_ext = attachment.name.lower()

            # 处理图片文件
            if file_ext.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                product_images.append({
                    'name': attachment.name,
                    'url': attachment.url
                })
                continue

            # 处理Excel文件
            if file_ext.endswith(('.xlsx', '.xls')):
                try:
                    contents = attachment.get_contents()
                    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))

                    if len(rows) > 1:
                        headers = rows[0]
                        headers_str = str(headers)

                        # 关键词数据表（检测特征列）
                        if '关键词' in headers_str and ('搜索量' in headers_str or '流量' in headers_str):
                            for row in rows[1:]:
                                if row and any(row):
                                    kw_dict = dict(zip(headers, row))
                                    keyword_data.append(kw_dict)
                        # 竞品数据表
                        elif 'ASIN' in headers_str and '评分数' in headers_str:
                            for row in rows[1:competitor_limit+1]:
                                if row and any(row):
                                    competitor_data.append(dict(zip(headers, row)))
                        # 评论数据表
                        elif 'ASIN' in headers_str and ('评论' in headers_str or '内容' in headers_str):
                            for row in rows[1:review_limit+1]:
                                if row and any(row):
                                    review_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取Excel文件 {attachment.name} 时出错：{str(e)}\n")

            # 处理CSV文件
            elif file_ext.endswith('.csv'):
                try:
                    contents = attachment.get_contents()
                    try:
                        text = contents.decode('utf-8-sig')
                    except:
                        text = contents.decode('gbk')

                    csv_reader = csv.DictReader(StringIO(text))
                    rows = list(csv_reader)

                    if len(rows) > 0:
                        headers = list(rows[0].keys())
                        if 'ASIN' in headers and any(h in headers for h in ['Rating', 'Body', 'Title', 'ReviewID']):
                            for row in rows[:review_limit]:
                                review_dict = {
                                    'ASIN': row.get('ASIN', ''),
                                    '标题': row.get('Title', ''),
                                    '内容': row.get('Body', ''),
                                    '星级': row.get('Rating', ''),
                                    '评论时间': row.get('Date', '')
                                }
                                review_data.append(review_dict)
                        elif 'ASIN' in headers and any(h in headers for h in ['商品标题', '评分数', '月销量']):
                            for row in rows[:competitor_limit]:
                                competitor_data.append(row)
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取CSV文件 {attachment.name} 时出错：{str(e)}\n")

        # 如果有数据，进行处理
        if competitor_data or review_data or reference_links or product_images or product_description or keyword_data:
            with poe.start_message() as msg:
                msg.write("## ✅ 数据收集完成\n\n")
                msg.write(f"📊 **竞品数据**：{len(competitor_data)} 个竞品\n")
                msg.write(f"💬 **评论数据**：{len(review_data)} 条评论\n")
                msg.write(f"🔗 **参考链接**：{len(reference_links)} 个链接\n")
                msg.write(f"🖼️ **产品图片**：{len(product_images)} 张图片\n")
                msg.write(f"📝 **产品描述**：{'有' if product_description else '无'}\n")
                msg.write(f"🔑 **关键词数据**：{len(keyword_data)} 个关键词\n\n")
                msg.write("⏳ 正在对数据进行AI预处理...\n\n")

            # AI预处理评论数据
            if review_data:
                review_text = "\n".join([
                    f"[评分{review.get('星级', 'N/A')}] {review.get('标题', '')} {review.get('内容', '')}"
                    for review in review_data
                ])

                review_prompt = f"""基于以下{len(review_data)}条亚马逊评论，生成结构化数据摘要：

{review_text}

请输出以下内容：
1. **评分分布**：各星级数量统计
2. **高频好评要点**：按提及次数排序，列出前10个
3. **高频差评要点**：按提及次数排序，列出前10个
4. **使用场景总结**：用户在什么场景下使用
5. **典型评论**：挑选最有代表性的好评3条、中评2条、差评3条（原文引用）

要求：量化、直观、全面保留关键信息。"""

                review_summary_response = poe.call("Gemini-3-Pro", review_prompt)

                with poe.start_message() as msg:
                    msg.write("### 📝 评论数据结构化摘要\n\n")
                    msg.write(review_summary_response.text)
                    msg.write("\n")

            # AI预处理竞品数据
            if competitor_data:
                comp_text = "\n".join([
                    f"{i+1}. 标题：{comp.get('商品标题', '')}\n   卖点：{comp.get('产品卖点', '')}"
                    for i, comp in enumerate(competitor_data)
                ])

                comp_prompt = f"""基于以下{len(competitor_data)}个竞品信息，生成结构化摘要：

{comp_text}

请输出：
1. **产品类型分析**：主要产品类型和特征
2. **高频卖点统计**：按出现频率排序，列出前15个常见卖点
3. **设计趋势**：材质、功能、风格等方面的趋势
4. **代表性产品**：挑选3个最有代表性的产品（包含标题和核心卖点）

要求：量化、直观、提炼关键信息。"""

                comp_summary_response = poe.call("Gemini-3-Pro", comp_prompt)

                with poe.start_message() as msg:
                    msg.write("### 🏆 竞品数据结构化摘要\n\n")
                    msg.write(comp_summary_response.text)
                    msg.write("\n")

            # 处理关键词数据（筛选+完整保留）
            if keyword_data:
                self.process_keyword_data(keyword_data, keyword_limit)

            # AI分析参考产品链接（使用web_search获取产品信息）
            if reference_links:
                self.analyze_reference_links(reference_links)

            # 显示其他收集的信息
            with poe.start_message() as msg:
                if product_description:
                    msg.write("### 📝 产品描述（与参考竞品的差异）\n\n")
                    msg.write(f"{product_description}\n\n")

                if product_images:
                    msg.write("### 🖼️ 产品图片（用于AI识别特征）\n\n")
                    for img in product_images:
                        msg.write(f"- 图片：{img['name']}\n")
                        msg.write(f"  - CDN链接：{img['url']}\n")
                    msg.write("\n")

                msg.write("【信息已收集】\n\n")
                msg.write("---\n\n")
                msg.write("💡 数据预处理完成！请回复「**继续**」开始基础分析（消费者洞察→使用流程→卖点提取）。\n")

        elif not is_first_run:
            with poe.start_message() as msg:
                msg.write("💡 请上传数据文件或粘贴链接。\n")

    def process_keyword_data(self, keyword_data, keyword_limit):
        """处理关键词数据：筛选+完整保留"""
        # 筛选强相关和高相关
        filtered_keywords = []
        for kw in keyword_data:
            relevance = kw.get('相关性档位', '')
            if relevance in ['强相关', '高相关']:
                filtered_keywords.append(kw)

        # 按周搜索量排序
        def get_search_volume(kw):
            vol = kw.get('周搜索量', 0)
            return int(vol) if vol else 0

        sorted_keywords = sorted(filtered_keywords, key=get_search_volume, reverse=True)

        # 限制数量
        final_keywords = sorted_keywords[:keyword_limit]

        # 生成摘要和完整列表
        with poe.start_message() as msg:
            msg.write("### 🔑 关键词数据（已筛选）\n\n")
            msg.write(f"**统计信息**\n")
            msg.write(f"- 原始词数：{len(keyword_data)} 个\n")
            msg.write(f"- 筛选后（强相关+高相关）：{len(filtered_keywords)} 个\n")
            msg.write(f"- 最终保留：{len(final_keywords)} 个\n\n")

            msg.write("**关键词完整列表（按搜索量排序）**\n\n")
            msg.write("| 序号 | 关键词 | 翻译 | 相关性 | 周搜索量 |\n")
            msg.write("|------|--------|------|--------|----------|\n")

            for i, kw in enumerate(final_keywords, 1):
                keyword = kw.get('关键词 (数据来源于西柚找词)', kw.get('关键词', ''))
                translation = kw.get('翻译', '')
                relevance = kw.get('相关性档位', '')
                volume = kw.get('周搜索量', 0)
                msg.write(f"| {i} | {keyword} | {translation} | {relevance} | {volume} |\n")

            msg.write("\n")

        # 新增：存储结构化 JSON，供后续精确切片
        keyword_json = []
        for kw in final_keywords:
            keyword_name = kw.get('关键词 (数据来源于西柚找词)', kw.get('关键词', ''))
            vol = kw.get('周搜索量', 0)
            keyword_json.append({
                "keyword": keyword_name,
                "translation": kw.get('翻译', ''),
                "relevance": kw.get('相关性档位', ''),
                "volume": int(vol) if vol else 0
            })
        self.store_summary("INFO_KEYWORDS", {
            "keywords": keyword_json,
            "stats": {
                "total_original": len(keyword_data),
                "total_filtered": len(filtered_keywords),
                "total_final": len(final_keywords)
            }
        })

    def analyze_reference_links(self, reference_links):
        """使用AI的web_search功能分析参考产品链接"""
        with poe.start_message() as msg:
            msg.write("### 🔗 参考产品分析\n\n")
            msg.write(f"正在通过网络搜索获取 {len(reference_links)} 个参考产品的详细信息...\n\n")

        links_text = "\n".join([f"- {link}" for link in reference_links])

        ref_prompt = f"""请访问以下亚马逊产品链接，获取并分析产品信息：

{links_text}

对于每个产品，请提取并整理以下信息：

## 参考产品1：[产品名称]
**链接**：[URL]

### 基本信息
- **完整标题**：[原文]
- **价格**：
- **评分**：
- **评论数**：

### 五点描述（Bullet Points）
1. [原文]
2. [原文]
3. [原文]
4. [原文]
5. [原文]

### 核心卖点分析
- **主打卖点**：[分析这个产品主打什么]
- **差异化特点**：[与同类产品的不同之处]
- **目标人群**：[面向什么用户]

### 主图分析
- **主图表达**：[主图展示了什么内容]
- **视觉风格**：[图片风格特点]

[如有多个链接，重复以上格式]

---

## 参考产品总结
- **共同特征**：[这些产品的共同点]
- **可借鉴之处**：[值得学习的地方]
- **可改进之处**：[可以做得更好的地方]

要求：
- 请务必访问链接获取真实信息，不要编造
- 尽可能提取完整的原文内容
- 分析要具体、有依据
"""

        # 使用Gemini-3-Pro的web_search功能
        ref_response = poe.call(
            "Gemini-3-Pro",
            poe.Message(text=ref_prompt, parameters={"web_search": True})
        )

        with poe.start_message() as msg:
            msg.write(ref_response.text)
            msg.write("\n\n")
            msg.write("【参考产品分析完成】\n\n")

    def handle_base_stage(self, user_input):
        """处理基础阶段（1-3）"""
        current_base_stage = self.detect_base_stage()

        # 用户说继续，进入下一阶段
        if "继续" in user_input or "下一步" in user_input:
            next_stage = current_base_stage + 1
            if next_stage >= len(self.base_stages):
                # 基础阶段全部完成，显示分支菜单
                self.show_branch_menu()
            else:
                self.execute_base_stage(next_stage)
        else:
            # 用户提供反馈，重新生成当前阶段
            if current_base_stage >= 0:
                self.execute_stage_with_feedback("base", current_base_stage, user_input)
            else:
                self.execute_base_stage(0)

    def detect_base_stage(self):
        """检测当前基础阶段"""
        chat_text = poe.default_chat.text

        if "【信息已收集】" not in chat_text:
            return -1

        for i in range(len(self.base_stages) - 1, -1, -1):
            if f"【基础阶段 {i+1} 完成】" in chat_text:
                return i

        return -1

    def execute_base_stage(self, stage_num):
        """执行基础阶段"""
        with poe.start_message() as msg:
            msg.write(f"# ⚙️ 基础阶段 {stage_num + 1}：{self.base_stages[stage_num]}\n\n")
            msg.write("AI分析中，请稍候...\n\n")

        prompt = self.build_base_prompt(stage_num)

        # 改造：不再传 poe.default_chat，所需上下文已在 prompt 中
        response = poe.call(
            "Gemini-3-Pro",
            prompt,
            output=poe.default_chat,
            adopt_current_bot_name=True
        )

        # 新增：生成并存储 JSON 摘要，供后续阶段精准引用
        summary_key = f"BASE_{stage_num + 1}"
        fields_desc = self._get_summary_fields_desc("base", stage_num)
        self.generate_and_store_summary(summary_key, response.text, fields_desc)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"【基础阶段 {stage_num + 1} 完成】{self.base_stages[stage_num]}\n\n")
            if stage_num < len(self.base_stages) - 1:
                msg.write("💡 请回复「**继续**」进入下一阶段，或提出调整要求。\n")
            else:
                # 基础阶段全部完成
                self.show_branch_menu()

    def show_branch_menu(self):
        """显示分支选择菜单"""
        chat_text = poe.default_chat.text

        with poe.start_message() as msg:
            msg.write("【基础分析完成】\n\n")
            msg.write("---\n\n")
            msg.write("## 🎯 请选择要生成的内容\n\n")
            msg.write("| 选项 | 内容 | 说明 |\n")
            msg.write("|------|------|------|\n")
            msg.write("| **1** | 主图+附图设计Brief | 7张图片的设计方案 |\n")
            msg.write("| **2** | A+页面设计Brief | A+页面模块设计方案 |\n")
            msg.write("| **3** | 标题+五点描述 | Listing文案优化 |\n")
            msg.write("| **4** | 全部生成 | 按顺序生成以上所有 |\n\n")

            # 检查已完成的分支
            completed = []
            if "【分支A完成】" in chat_text:
                completed.append("主图+附图")
            if "【分支B完成】" in chat_text:
                completed.append("A+页面")
            if "【分支C完成】" in chat_text:
                completed.append("标题五点")

            if completed:
                msg.write(f"✅ 已完成：{', '.join(completed)}\n\n")

            # 检查关键词数据
            if "🔑 关键词数据" not in chat_text:
                msg.write("⚠️ 提示：如需生成「标题+五点」，请先上传关键词数据表格。\n\n")

            msg.write("请输入数字（1-4）或名称选择：\n")

    def handle_branch(self, user_input):
        """处理分支选择和执行"""
        chat_text = poe.default_chat.text

        # 检查是否正在某个分支中
        current_branch, current_stage = self.detect_branch_stage()

        # 如果在分支中
        if current_branch:
            if "继续" in user_input or "下一步" in user_input:
                self.continue_branch(current_branch, current_stage)
            else:
                # 用户反馈，重新生成
                self.execute_stage_with_feedback(current_branch, current_stage, user_input)
            return

        # 检查是否需要上传关键词（选择了分支C但没有关键词数据）
        # 注意：如果已有关键词数据（🔑标记），则不再要求上传
        if "【等待关键词数据】" in chat_text and "🔑 关键词数据" not in chat_text:
            self.handle_keyword_upload(user_input)
            return

        # 解析用户选择
        choice = self.parse_branch_choice(user_input)

        if choice == "A":
            self.start_branch_a()
        elif choice == "B":
            self.start_branch_b()
        elif choice == "C":
            # 检查是否有关键词数据
            if "🔑 关键词数据" not in chat_text:
                self.request_keyword_upload()
            else:
                self.start_branch_c()
        elif choice == "ALL":
            # 标记全部生成模式
            with poe.start_message() as msg:
                msg.write("【全部生成模式】\n\n")
            self.start_branch_a()  # 先从A开始，完成后自动进入下一个
        else:
            # 无法识别，显示菜单
            self.show_branch_menu()

    def parse_branch_choice(self, user_input):
        """解析用户的分支选择"""
        user_input = user_input.lower()

        if "1" in user_input or "主图" in user_input or "附图" in user_input:
            return "A"
        elif "2" in user_input or "a+" in user_input or "A+" in user_input:
            return "B"
        elif "3" in user_input or "���题" in user_input or "五点" in user_input:
            return "C"
        elif "4" in user_input or "全部" in user_input or "都要" in user_input:
            return "ALL"

        return None

    def request_keyword_upload(self):
        """请求上传关键词数据"""
        with poe.start_message() as msg:
            msg.write("## 🔑 需要关键词数据\n\n")
            msg.write("生成「标题+五点描述」需要关键词数据表格。\n\n")
            msg.write("请上传包含以下列的Excel表格：\n")
            msg.write("- 关键词\n")
            msg.write("- 翻译（可选）\n")
            msg.write("- 相关性档位（强相关/高相关等）\n")
            msg.write("- 周搜索量\n\n")
            msg.write("【等待关键词数据】\n")

    def handle_keyword_upload(self, user_input):
        """处理关键词数据上传"""
        keyword_limit = self.get_limits()[2]
        keyword_data = []

        for attachment in poe.query.attachments:
            file_ext = attachment.name.lower()
            if file_ext.endswith(('.xlsx', '.xls')):
                try:
                    contents = attachment.get_contents()
                    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))

                    if len(rows) > 1:
                        headers = rows[0]
                        for row in rows[1:]:
                            if row and any(row):
                                keyword_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取文件出错：{str(e)}\n")

        if keyword_data:
            self.process_keyword_data(keyword_data, keyword_limit)
            # 检查分支C是否已经开始
            chat_text = poe.default_chat.text
            if "【分支C开始】" in chat_text and "【分支C完成】" not in chat_text:
                # 分支C已在进行中，继续当前阶段
                current_branch, current_stage = self.detect_branch_stage()
                if current_branch == "C":
                    self.continue_branch("C", current_stage)
                else:
                    # 某些异常情况，重新开始分支C
                    self.start_branch_c()
            else:
                # 分支C还没开始，开始它
                self.start_branch_c()
        else:
            # 检查是否实际上已经有关键词数据（用户可能只是发了不相关的消息）
            chat_text = poe.default_chat.text
            if "🔑 关键词数据" in chat_text:
                # 已有关键词数据，可能用户只是在聊天
                with poe.start_message() as msg:
                    msg.write("💡 关键词数据已存在。请选择要执行的操作，或回复「继续」进行下一步。\n")
            else:
                with poe.start_message() as msg:
                    msg.write("⚠️ 未检测到关键词数据，请上传Excel表格。\n")

    def detect_branch_stage(self):
        """检测当前所在的分支和阶段（检测最后一个活动的分支）"""
        chat_text = poe.default_chat.text

        # 找出所有活动的分支及其最后出现的位置
        active_branches = []

        # 检查分支C（优先检查，因为在全部生成模式下C是最后的）
        if "【分支C开始】" in chat_text and "【分支C完成】" not in chat_text:
            pos = chat_text.rfind("【分支C开始】")
            active_branches.append(("C", pos, self.branch_c_stages))

        # 检查分支B
        if "【分支B开始】" in chat_text and "【分支B完成】" not in chat_text:
            pos = chat_text.rfind("【分支B开始】")
            active_branches.append(("B", pos, self.branch_b_stages))

        # 检查分支A
        if "【分支A开始】" in chat_text and "【分支A完成】" not in chat_text:
            pos = chat_text.rfind("【分支A开始】")
            active_branches.append(("A", pos, self.branch_a_stages))

        if not active_branches:
            return None, -1

        # 选择位置最靠后的分支（最新开始的分支）
        active_branches.sort(key=lambda x: x[1], reverse=True)
        branch_name, _, stages = active_branches[0]

        # 找到该分支完成的最后一个阶段
        for i in range(len(stages) - 1, -1, -1):
            if f"【分支{branch_name}-阶段 {i+1} 完成】" in chat_text:
                return branch_name, i

        return branch_name, -1

    def start_branch_a(self):
        """开始分支A：主副图"""
        with poe.start_message() as msg:
            msg.write("## 🎨 开始生成：主图+附图设计Brief\n\n")
            msg.write("【分支A开始】\n\n")

        self.execute_branch_stage("A", 0)

    def start_branch_b(self):
        """开始分支B：A+页面"""
        with poe.start_message() as msg:
            msg.write("## 📄 开始生成：A+页面设计Brief\n\n")
            msg.write("【分支B开始】\n\n")

        self.execute_branch_stage("B", 0)

    def start_branch_c(self):
        """开始分支C：标题五点"""
        with poe.start_message() as msg:
            msg.write("## 📝 开始生成：标题+五点描述\n\n")
            msg.write("【分支C开始】\n\n")

        self.execute_branch_stage("C", 0)

    def continue_branch(self, branch, current_stage):
        """继续当前分支的下一阶段"""
        next_stage = current_stage + 1

        if branch == "A":
            stages = self.branch_a_stages
        elif branch == "B":
            stages = self.branch_b_stages
        else:
            stages = self.branch_c_stages

        if next_stage >= len(stages):
            self.complete_branch(branch)
        else:
            self.execute_branch_stage(branch, next_stage)

    def execute_branch_stage(self, branch, stage_num):
        """执行分支阶段"""
        if branch == "A":
            stages = self.branch_a_stages
        elif branch == "B":
            stages = self.branch_b_stages
        else:
            stages = self.branch_c_stages

        with poe.start_message() as msg:
            msg.write(f"# ⚙️ 分支{branch}-阶段 {stage_num + 1}：{stages[stage_num]}\n\n")
            msg.write("AI分析中，请稍候...\n\n")

        prompt = self.build_branch_prompt(branch, stage_num)

        # 改造：不再传 poe.default_chat，所需上下文已在 prompt 中
        response = poe.call(
            "Gemini-3-Pro",
            prompt,
            output=poe.default_chat,
            adopt_current_bot_name=True
        )

        # 新增：生成并存储 JSON 摘要
        summary_key = f"BRANCH_{branch}_{stage_num + 1}"
        fields_desc = self._get_summary_fields_desc(branch, stage_num)
        self.generate_and_store_summary(summary_key, response.text, fields_desc)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"【分支{branch}-阶段 {stage_num + 1} 完成】{stages[stage_num]}\n\n")

            if stage_num < len(stages) - 1:
                msg.write("💡 请回复「**继续**」进入下一阶段，或提出调整要求。\n")
            else:
                self.complete_branch(branch)

    def complete_branch(self, branch):
        """完成分支"""
        chat_text = poe.default_chat.text
        is_all_mode = "【全部生成模式】" in chat_text

        # 先输出完成标记
        with poe.start_message() as msg:
            msg.write(f"【分支{branch}完成】\n\n")

            if branch == "A":
                msg.write("✅ 主图+附图设计Brief 已生成完成！\n\n")
            elif branch == "B":
                msg.write("✅ A+页面设计Brief 已生成完成！\n\n")
            else:
                msg.write("✅ 标题+五点描述 已生成完成！\n\n")

        # 检查是否是"全部生成"模式，继续下一个分支（在with块外调用）
        if is_all_mode:
            if branch == "A" and "【分支B完成】" not in chat_text:
                with poe.start_message() as msg:
                    msg.write("正在继续生成下一项...\n\n")
                self.start_branch_b()
                return
            elif branch == "B" and "【分支C完成】" not in chat_text:
                if "🔑 关键词数据" in chat_text:
                    with poe.start_message() as msg:
                        msg.write("正在继续生成下一项...\n\n")
                    self.start_branch_c()
                    return
                else:
                    with poe.start_message() as msg:
                        msg.write("⚠️ 缺少关键词数据，无法生成标题五点。\n\n")

        # 显示菜单
        with poe.start_message() as msg:
            msg.write("---\n\n")
            msg.write("您可以：\n")
            msg.write("- 输入数字选择生成其他内容\n")
            msg.write("- 对当前结果提出修改意见\n\n")

        self.show_branch_menu()

    def execute_stage_with_feedback(self, stage_type, stage_num, user_feedback):
        """根据用户反馈重新执行阶段"""
        if stage_type == "base":
            stage_name = self.base_stages[stage_num]
            prompt = self.build_base_prompt(stage_num)
            marker = f"【基础阶段 {stage_num + 1} 完成】{stage_name}"
        else:
            if stage_type == "A":
                stage_name = self.branch_a_stages[stage_num]
            elif stage_type == "B":
                stage_name = self.branch_b_stages[stage_num]
            else:
                stage_name = self.branch_c_stages[stage_num]
            prompt = self.build_branch_prompt(stage_type, stage_num)
            marker = f"【分支{stage_type}-阶段 {stage_num + 1} 完成】{stage_name}"

        with poe.start_message() as msg:
            msg.write(f"## 🔄 根据您的反馈重新生成\n\n")
            msg.write(f"**您的要求**：{user_feedback}\n\n")
            msg.write("AI正在重新生成，请稍候...\n\n")

        enhanced_prompt = f"""{prompt}

---

**用户特别要求：**
{user_feedback}

请严格遵守用户的要求重新生成内容。特别注意：
- 保留所有详细信息，不要精简
- 如果用户要求增加内容，务必添加
- 如果用户要求修改某部分，只修改该部分，其他部分保持原样"""

        # 改造：不再传 poe.default_chat
        response = poe.call(
            "Gemini-3-Pro",
            enhanced_prompt,
            output=poe.default_chat,
            adopt_current_bot_name=True
        )

        # 新增：重新生成摘要（覆盖旧版本）
        if stage_type == "base":
            summary_key = f"BASE_{stage_num + 1}"
        else:
            summary_key = f"BRANCH_{stage_type}_{stage_num + 1}"
        fields_desc = self._get_summary_fields_desc(stage_type, stage_num)
        self.generate_and_store_summary(summary_key, response.text, fields_desc)

        with poe.start_message() as msg:
            msg.write("\n\n---\n\n")
            msg.write(f"{marker}\n\n")
            msg.write("💡 请回复「**继续**」进入下一阶段，或提出调整要求。\n")

    # ============ 基础上下文 ============

    def get_base_context(self):
        """获取所有基础上下文信息（不含关键词，关键词仅在分支C中按需注入）"""
        reference_analysis = self.extract_reference_analysis()
        product_description = self.extract_section("📝 产品描述（与参考竞品的差异）")
        product_images = self.extract_section("🖼️ 产品图片")
        review_summary = self.extract_section("📝 评论数据结构化摘要")
        competitor_summary = self.extract_section("🏆 竞品数据结构化摘要")

        context_parts = ["# 基础信息（所有阶段均可参考）\n"]

        # 参考产品分析（最重要，放在最前面）
        if reference_analysis:
            context_parts.append("## 参考产品分析（本产品的主要依据，务必参考）\n")
            context_parts.append(f"{reference_analysis}\n")

        if product_description:
            context_parts.append("## 产品描述（与参考竞品的差异）\n")
            context_parts.append(f"{product_description}\n")

        if product_images:
            context_parts.append("## 产品图片\n")
            context_parts.append(f"{product_images}\n")

        if review_summary:
            context_parts.append("## 评论数据摘要\n")
            context_parts.append(f"{review_summary}\n")

        if competitor_summary:
            context_parts.append("## 竞品数据摘要\n")
            context_parts.append(f"{competitor_summary}\n")

        # 关键词数据已移除：仅在分支C（标题五点）中按需注入，
        # 避免在不需要关键词的阶段（消费者洞察、主副图、A+页面等）浪费 Token

        context_parts.append("---\n\n")

        return "\n".join(context_parts)

    def extract_reference_analysis(self):
        """提取参考产品分析内容（从🔗标记到【参考产品分析完成】）"""
        chat_text = poe.default_chat.text

        start_marker = "### 🔗 参考产品分析"
        end_marker = "【参考产品分析完成】"

        if start_marker not in chat_text:
            return ""

        start_pos = chat_text.find(start_marker)
        end_pos = chat_text.find(end_marker)

        if start_pos >= 0 and end_pos > start_pos:
            content = chat_text[start_pos:end_pos]
            # 清理开头的标记
            content = content.replace(start_marker, "").strip()
            # 清理"正在通过网络搜索..."提示
            if "正在通过网络搜索" in content:
                lines = content.split("\n")
                content = "\n".join([l for l in lines if "正在通过网络搜索" not in l])
            return content.strip()[:8000]  # 参考产品分析可能较长，给更多空间

        return ""

    def extract_section(self, section_marker):
        """从对话历史中提取特定段落"""
        chat_text = poe.default_chat.text

        if section_marker not in chat_text:
            return ""

        parts = chat_text.split(section_marker)
        if len(parts) > 1:
            content = parts[1]
            end_markers = ["###", "【", "---", "##"]
            min_pos = len(content)
            for marker in end_markers:
                pos = content.find(marker, 50)
                if pos > 0 and pos < min_pos:
                    min_pos = pos
            return content[:min_pos].strip()[:3000]

        return ""

    def extract_stage_output(self, stage_type, stage_num):
        """提取指定阶段的输出"""
        chat_text = poe.default_chat.text

        if stage_type == "base":
            marker = f"【基础阶段 {stage_num + 1} 完成】"
        else:
            marker = f"【分支{stage_type}-阶段 {stage_num + 1} 完成】"

        if marker not in chat_text:
            return "[未找到阶段输出]"

        marker_pos = chat_text.rfind(marker)
        before_marker = chat_text[:marker_pos]

        # 找到阶段开始位置
        if stage_type == "base":
            if stage_num == 0:
                start_marker = "【信息已收集】"
            else:
                start_marker = f"【基础阶段 {stage_num} 完成】"
        else:
            if stage_num == 0:
                start_marker = f"【分支{stage_type}开始】"
            else:
                start_marker = f"【分支{stage_type}-阶段 {stage_num} 完成】"

        if start_marker in before_marker:
            start_pos = before_marker.rfind(start_marker)
        else:
            start_pos = 0

        content = chat_text[start_pos:marker_pos]

        # 清理内容
        lines = content.split('\n')
        clean_lines = [line for line in lines if not any(x in line for x in ['【', '💡', '请回复', 'AI分析中', '# ⚙️'])]

        result = '\n'.join(clean_lines).strip()
        return result[:4000] if len(result) > 100 else "[未找到有效阶段输出]"

    def get_previous_branch_outputs(self):
        """获取之前分支的输出摘要（用于可选参考）"""
        chat_text = poe.default_chat.text
        outputs = {}

        if "【分支A完成】" in chat_text:
            summary = self.get_summary("BRANCH_A_3")  # Brief 生成阶段的摘要
            if summary and "raw_excerpt" not in summary:
                outputs["主图附图Brief"] = json.dumps(summary, ensure_ascii=False, indent=2)
            else:
                # 降级到旧式提取
                outputs["主图附图Brief"] = self.extract_section("主图+附图设计Brief")

        if "【分支B完成】" in chat_text:
            summary = self.get_summary("BRANCH_B_2")  # Brief 生成阶段的摘要
            if summary and "raw_excerpt" not in summary:
                outputs["A+页面Brief"] = json.dumps(summary, ensure_ascii=False, indent=2)
            else:
                outputs["A+页面Brief"] = self.extract_section("A+页面设计Brief")

        return outputs

    # ============ Prompt构建 ============

    def build_base_prompt(self, stage_num):
        """构建基础阶段的prompt"""
        base_context = self.get_base_context()

        if stage_num == 0:
            return f"""{base_context}

你是一位资深的亚马逊运营专家和消费者洞察分析师。

请基于上述基础信息，进行深度的消费者洞察分析。

请按照以下结构输出分析结果（使用中文）：

## 1. 消费者画像
- 主要用户群体特征
- 使用场景
- 购买动机

## 2. 购物决策因素
- 最重要的3-5个决策因素
- 每个因素的重要程度和原因

## 3. 核心痛点
- 用户最关心的痛点（按优先级排序）
- 每个痛点的具体表现

## 4. 期望与需求
- 用户明确提出的需求
- 隐含的期望

要求：
- 直观直白，避免理想化表达
- 基于数据事实，不要主观臆测
- 优先级明确，有理有据
"""

        elif stage_num == 1:
            consumer_insights = self._get_stage_summary_text("base", 0)
            return f"""{base_context}

你是一位产品体验设计专家。

基于上述基础信息和消费者洞察摘要，请完整复现产品的使用流程：

**消费者洞察摘要：**
{consumer_insights}

请按照时间顺序，详细描述产品使用的完整过程，包括：

## 使用前
- 用户的准备工作
- 需要了解的信息
- 开箱/安装步骤

## 使用中
- 具体使用步骤（分步骤详细说明）
- 每个步骤的操作要点
- 可能遇到的场景

## 使用后
- 收纳/维护
- 清洁保养
- 长期使用注意事项

要求：
- 力求完整，不遗漏任何步骤
- 代入消费者视角，真实还原使用场景
- 每个步骤都要具体、可操作
- 直白表达，不要过于理想化
"""

        elif stage_num == 2:
            usage_flow = self._get_stage_summary_text("base", 1)
            return f"""{base_context}

你是一位产品竞争分析专家。

基于上述基础信息和使用流程摘要，提取产品在每个使用步骤的卖点：

**使用流程摘要：**
{usage_flow}

请按照使用流程的每个步骤，分析：
1. 该步骤涉及的产品功能或特性
2. 与竞品对比，可能的优势是什么
3. 可以表达的卖点

输出格式：

## 使用前
### 步骤1：[步骤名称]
- **涉及功能**：
- **竞品对比**：
- **可表达卖点**：

[重复其他步骤]

## 使用中
[同上格式]

## 使用后
[同上格式]

## 卖点汇总（按重要性排序）
1. [最重要的卖点]
2. [次要卖点]
...

要求：
- 卖点要具体、可量化
- 对比要客观、有依据
- 避免空洞的形容词
- 直白表达
"""

        return "[未定义阶段]"

    def build_branch_prompt(self, branch, stage_num):
        """构建分支阶段的prompt"""
        base_context = self.get_base_context()
        selling_points = self._get_stage_summary_text("base", 2)  # 卖点摘要
        previous_outputs = self.get_previous_branch_outputs()

        # 可选参考部分
        reference_section = ""
        if previous_outputs:
            reference_section = "\n## 可选参考（之前生成的内容）\n"
            for name, content in previous_outputs.items():
                if content:
                    reference_section += f"\n### {name}\n{content[:1500]}...\n"

        if branch == "A":
            return self.build_branch_a_prompt(stage_num, base_context, selling_points, reference_section)
        elif branch == "B":
            return self.build_branch_b_prompt(stage_num, base_context, selling_points, reference_section)
        else:
            return self.build_branch_c_prompt(stage_num, base_context, selling_points, reference_section)

    def build_branch_a_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支A：主副图"""
        if stage_num == 0:
            # A1：卖点选择与排序
            consumer_insights = self._get_stage_summary_text("base", 0)
            return f"""{base_context}
{reference_section}

你是一位亚马逊Listing优化专家。

基于基础分析结果，请为亚马逊主图+6张附图（共7张）确定表达内容和顺序。

**消费者洞察摘要：**
{consumer_insights}

**卖点摘要：**
{selling_points}

请输出：

## 图片规划

### 主图（第1张）
- **表达卖点**：[最核心的卖点]
- **选择理由**：[为什么这个卖点最重要]

### 附图1（第2张）
- **表达卖点**：
- **选择理由**：

### 附图2（第3张）
- **表达卖点**：
- **选择理由**：

### 附图3（第4张）
- **表达卖点**：
- **选择理由**：

### 附图4（第5张）
- **表达卖点**：
- **选择理由**：

### 附图5（第6张）
- **表达卖点**：
- **选择理由**：

### 附图6（第7张）
- **表达卖点**：
- **选择理由**：

## 排序逻辑说明
[解释为什么按这个顺序排列]

要求：
- 结合消费者最关心的点进行优先排序
- 考虑视觉呈现的逻辑性和连贯性
- 确保覆盖核心购买决策因素
"""

        elif stage_num == 1:
            # A2：卖点表达方式设计
            image_plan = self._get_stage_summary_text("A", 0)
            return f"""{base_context}

你是一位创意设计策略专家，精通各行业的视觉表达方式。

基于图片规划摘要，为每张图片设计专业的表达方式和AI绘图prompt。

**图片规划摘要：**
{image_plan}

对于每张图片，请：
1. 研究该卖点在同类产品中的常见表达方式
2. 思考其他行业/品牌如何专业地表达类似卖点
3. 提供具体的设计建议
4. 生成详细的AI绘图prompt

输出格式：

## 主图（第1张）

### 卖点回顾
[重复卖点]

### 表达方式研究
- **同类产品常见表达**：
- **跨行业优秀案例**：
- **设计建议**：

### AI绘图Prompt
```
[详细的prompt]
```

[对其余6张附图重复以上格式]
"""

        elif stage_num == 2:
            # A3：设计Brief生成
            image_plan = self._get_stage_summary_text("A", 0)
            design_approach = self._get_stage_summary_text("A", 1)
            consumer_insights = self._get_stage_summary_text("base", 0)

            return f"""{base_context}

你是一位专业的设计项目经理。

请生成一份完整的、适合外包设计公司使用的主图+附图设计Brief文档。

**消费者洞察摘要：**
{consumer_insights}

**图片规划摘要：**
{image_plan}

**设计方案摘要：**
{design_approach}

请生成一份结构清晰的Markdown格式设计Brief，包含：

# 亚马逊Listing图片设计Brief

## 1. 项目概述
- 产品类型
- 设计目标
- 交付要求（7张图片：1张主图+6张附图）

## 2. 消费者洞察总结

## 3. 图片设计需求详解

### 3.1 主图（第1张）
- **核心卖点**：
- **图片文案**：
  - **主标题**：[3-6个英文单词]
  - **副标题**：[可选]
- **构图设计**：
  - **画面布局**：
  - **拍摄角度**：
  - **产品状态**：
  - **背景设置**：
- **文字排版布局**：
  - **主标题位置**：
  - **���标题样式**：
- **视觉元素布局**：
- **色彩方案**：
- **AI绘图Prompt**：

[对其他6张图重复以上格式]

## 4. 设计规范
## 5. 验收标准清单

要求：
- 每张图片都必须有明确的英文文案
- 构图设计必须详细具体
- 色彩方案必须明确
"""

        return "[未定义阶段]"

    def build_branch_b_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支B：A+页面"""
        if stage_num == 0:
            # B1：A+页面模块规划
            consumer_insights = self._get_stage_summary_text("base", 0)
            usage_flow = self._get_stage_summary_text("base", 1)

            return f"""{base_context}
{reference_section}

你是一位亚马逊A+页面设计专家。

A+页面是品牌展示和深度说服的核心工具，需要通过多个模块组合来全面呈现产品价值。

基于基础分析摘要，为A+页面规划模块布局：

**消费者洞察摘要：**
{consumer_insights}

**使用流程摘要：**
{usage_flow}

**卖点摘要：**
{selling_points}

请规划A+页面的模块设计（建议5个模块，可根据需要增减）：

## A+页面模块规划

### 模块1：[模块类型]
- **模块名称**：
- **核心目标**：
- **表达内容**：
- **布局建议**：

[重复其他模块]

## 模块排序逻辑
[解释为什么按这个顺序排列]
"""

        elif stage_num == 1:
            # B2：A+页面设计Brief
            module_plan = self._get_stage_summary_text("B", 0)

            return f"""{base_context}

你是一位专业的A+页面设计项目经理。

请基于模块规划摘要，生成详细的A+页面设计Brief文档。

**模块规划摘要：**
{module_plan}

请生成：

# 亚马逊A+页面设计Brief

## 1. A+页面概述
- 设计目标
- 整体风格
- 模块总数

## 2. 模块设计需求详解

### 2.1 模块1：[模块名称]
- **核心目标**：
- **模块文案**：
  - **主标题**：[英文]
  - **副标题**：
  - **正文**：
- **模块布局**：
- **文字排版**：
- **视觉元素**：
- **色彩方案**：
- **图片需求**：
- **AI绘图Prompt**：

[重复其他模块]

## 3. 设计规范
## 4. 验收标准
"""

        return "[未定义阶段]"

    def build_branch_c_prompt(self, stage_num, base_context, selling_points, reference_section):
        """分支C：标题五点"""
        if stage_num == 0:
            # C1：关键词策略分析 —— 使用完整关键词列表
            keyword_data = self._get_keyword_data_for_prompt()  # 全量
            return f"""{base_context}
{reference_section}

你是一位亚马逊SEO和关键词策略专家。

基于以下关键词数据和卖点分析，制定关键词策略：

{keyword_data}

**卖点摘要：**
{selling_points}

请输出：

## 关键词策略分析

### 1. 核心主词（必须埋入标题）
| 关键词 | 周搜索量 | 选择理由 |
|--------|----------|----------|
| ... | ... | ... |

（建议3-5个核心主词）

### 2. 重要长尾词（标题或五点使用）
| 关键词 | 周搜索量 | 建议位置 |
|--------|----------|----------|
| ... | ... | 标题/五点第X条 |

（建议10-15个长尾词）

### 3. 五点专用词（按卖点分组）
| 卖点主题 | 相关关键词 | 建议融入方式 |
|----------|------------|--------------|
| 卖点1：XX | kw1, kw2 | ... |
| ... | ... | ... |

### 4. 关键词布局建议
- **标题**：[建议包含哪些词，大致顺序]
- **五点1**：[建议包含哪些词]
- **五点2**：[建议包含哪些词]
- **五点3**：[建议包含哪些词]
- **五点4**：[建议包含哪些词]
- **五点5**：[建议包含哪些词]
"""

        elif stage_num == 1:
            # C2：标题生成
            keyword_strategy = self._get_stage_summary_text("C", 0)
            consumer_insights = self._get_stage_summary_text("base", 0)

            return f"""{base_context}

你是一位亚马逊Listing标题优化专家。

基于关键词策略摘要和产品卖点，生成标题方案：

**关键词策略摘要：**
{keyword_strategy}

**消费者洞察摘要：**
{consumer_insights}

**卖点摘要：**
{selling_points}

请输出：

## 标题生成

### 标题要求
- 字符限制：200字符以内
- 必须包含核心主词
- 突出最重要的1-2个卖点
- 符合亚马逊规范（无促销语、无特殊符号）

### 标题方案1（推荐）
**标题**：[完整标题]
**字符数**：[X字符]
**包含关键词**：[列出包含的核心词]
**突出卖点**：[列出突出的卖点]
**设计思路**：[解释为什么这样写]

### 标题方案2（备选）
[同上格式]

### 标题方案3（备选）
[同上格式]

### 方案对比
| 方案 | 优点 | 缺点 | 推荐度 |
|------|------|------|--------|
| 1 | ... | ... | ⭐⭐⭐⭐⭐ |
| 2 | ... | ... | ... |
| 3 | ... | ... | ... |
"""

        elif stage_num == 2:
            # C3：五点描述生成
            keyword_strategy = self._get_stage_summary_text("C", 0)
            title_output = self._get_stage_summary_text("C", 1)
            consumer_insights = self._get_stage_summary_text("base", 0)

            return f"""{base_context}

你是一位亚马逊Listing五点描述优化专家。

基于关键词策略摘要、标题和产品卖点，生成五点描述：

**关键词策略摘要：**
{keyword_strategy}

**标题方案摘要：**
{title_output}

**消费者洞察摘要：**
{consumer_insights}

**卖点摘要：**
{selling_points}

请输出：

## 五点描述（Bullet Points）

### 五点要求
- 每条500字符以内
- 每条聚焦一个核心卖点
- 自然融入相关关键词
- 结构：卖点标题（大写） + 详细描述
- 符合亚马逊规范

### 五点1
**主题**：[卖点主题]
**包含关键词**：[列出融入的关键词]
**内容**：
```
[完整五点内容，含英文大写标题]
```
**字符数**：[X字符]

### 五点2
[同上格式]

### 五点3
[同上格式]

### 五点4
[同上格式]

### 五点5
[同上格式]

### 关键词覆盖检查
| 关键词 | 出现位置 |
|--------|----------|
| kw1 | 标题、五点1 |
| kw2 | 五点2 |
| ... | ... |
"""

        return "[未定义阶段]"


if __name__ == "__main__":
    bot = ListingDesigner()
    bot.run()