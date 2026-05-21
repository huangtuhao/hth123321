# poe: name=Listing-Designer

import openpyxl
import csv
import re
import json
import time
from io import BytesIO, StringIO
from fastapi_poe.types import SettingsResponse

poe.update_settings(SettingsResponse(
    introduction_message="""# 🚀 亚马逊Listing设计助手

欢迎使用！我将帮助您完成Listing的完整设计流程。

## 📋 第一步：信息收集

请在**一条消息**中上传以下信息：
1. **竞品数据Excel/CSV表格**（含ASIN、标题、五点等）
2. **评论数据Excel/CSV表格**（竞品评论）
3. **参考产品链接**（完整URL，作为本产品的主要依据）
4. **产品图片**（可选，用于AI识别产品外观特征）
5. **产品描述**（可选，说明与参考竞品的不同之处）
6. **关键词数据Excel**（可选，也可在生成标题五点时再上传）
7. **品牌名称**（可选，注明"品牌：XXX"即可）

## 💡 使用指南

**交互命令（任意阶段可用）：**
- 「**继续**」→ 进入下一阶段
- 「**重新生成**」→ 重做当前阶段（可附带修改意见）
- 「**进度**」→ 查看流程进度和各阶段统计
- 「**AI：模型名**」→ 切换AI模型（如"AI：Claude-Sonnet-4.6"）
- 「**参数：评论500，竞品50，关键词80**」→ 调整数据读取量

**讨论模式：**
每个阶段完成后，可以直接输入问题或修改意见进行讨论。讨论不消耗额外token——AI只看到当前阶段的结论和您的问题。

**流程概览：**
基础分析（3步）→ 选择生成内容（主图Brief / A+ Brief / 标题五点 / 全部）""",
))


# ============ JSON 解析工具 ============

def parse_json_from_response(response_text, required_fields=None):
    """从 LLM 响应中提取并解析 JSON，带容错处理。"""
    text = re.sub(r'\*Thinking\.\.\.\*\n\n(?:> [^\n]*\n?)*', '', response_text)
    text = re.sub(r'\n*---\n+(?:Learn more|Related searches):\n[\s\S]*$', '', text)

    match = re.search(r'```json\s*([\[{].*?[\]}])\s*```', text, re.DOTALL)
    if match:
        json_str = match.group(1).strip()
    else:
        match = re.search(r'[\[{].*[\]}]', text, re.DOTALL)
        if match:
            json_str = match.group(0).strip()
        else:
            raise Exception(f"无法在响应中找到 JSON:\n{response_text[:500]}")

    try:
        parsed = json.loads(json_str)
    except json.JSONDecodeError as e:
        cleaned = re.sub(r',\s*([}\]])', r'\1', json_str)
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            raise Exception(f"JSON 解析失败: {e}\n原始文本: {json_str[:500]}")

    if required_fields:
        missing = set(required_fields) - set(parsed.keys())
        if missing:
            raise Exception(f"JSON 缺少必需字段: {', '.join(missing)}")
    return parsed


# ============ 全局状态管理 ============

class GlobalState:
    """基于聊天历史的结构化状态管理。

    所有状态以 JSON 存储在聊天消息中。AI 调用不传对话历史，
    只通过 Python 从状态中精确提取需要的字段。
    """

    @staticmethod
    def store(key, data):
        json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        with poe.start_message() as msg:
            msg.write(f"【STATE:{key}】{json_str}【/STATE】\n")

    @staticmethod
    def get(key):
        chat_text = poe.default_chat.text
        pattern = f'【STATE:{re.escape(key)}】(.*?)【/STATE】'
        matches = re.findall(pattern, chat_text, re.DOTALL)
        if not matches:
            return None
        try:
            return json.loads(matches[-1].strip())
        except json.JSONDecodeError:
            return None

    @staticmethod
    def has(key):
        return GlobalState.get(key) is not None

    @staticmethod
    def require(*keys):
        """校验关键 key 是否全部存在。丢失则抛出明确错误。
        返回 dict: {key: data}。
        """
        missing = []
        results = {}
        for key in keys:
            data = GlobalState.get(key)
            if data is None:
                missing.append(key)
            else:
                results[key] = data
        if missing:
            labels = {
                "PRODUCT_INFO": "产品信息", "REVIEW_EXTRACTION": "评论提取",
                "COMPETITOR_EXTRACTION": "竞品提取", "REFERENCE_ANALYSIS": "参考产品",
                "CONSUMER_INSIGHTS": "消费者洞察", "USAGE_FLOW": "使用流程",
                "SELLING_POINTS": "卖点列表", "KEYWORDS": "关键词",
                "PROGRESS": "进度状态",
            }
            names = [labels.get(k, k) for k in missing]
            raise poe.BotError(
                f"⚠️ 关键数据丢失：{', '.join(names)}。\n"
                f"可能原因：对话过长导致历史被截断。\n"
                f"建议：开一个新对话，重新上传数据文件。"
            )
        return results

    @staticmethod
    def check_degraded(*keys):
        """检查指定 key 中是否有降级数据（_raw_text）。
        返回降级的 key 列表。
        """
        degraded = []
        for key in keys:
            data = GlobalState.get(key)
            if data and "_raw_text" in data:
                degraded.append(key)
        return degraded

    @staticmethod
    def get_progress():
        return GlobalState.get("PROGRESS") or {
            "base_completed": 0,
            "current_branch": None,
            "branch_stages": {},
            "completed_branches": [],
            "all_mode": False
        }

    @staticmethod
    def set_progress(progress):
        GlobalState.store("PROGRESS", progress)


# ============ Listing 设计器 ============

class ListingDesigner:
    EXTRACTION_MODEL = "Gemini-3-Flash"
    DEFAULT_MODEL = "Gemini-3.1-Pro"

    def __init__(self):
        self.base_stages = ["消费者洞察分析", "使用流程复现", "卖点提取"]
        self.branch_a_stages = ["卖点选择与排序", "卖点表达方式设计", "主图+附图设计Brief生成"]
        self.branch_b_stages = ["A+页面模块规划", "A+页面设计Brief生成"]
        self.branch_c_stages = ["关键词策略分析", "标题生成", "五点描述生成"]

    def run(self):
        user_input = poe.query.text.strip()

        # --- 全局命令（任意阶段可用）---
        if user_input.startswith("参数：") or user_input.startswith("参数:"):
            self._handle_parameters(user_input)
            return

        if user_input in ("进度", "状态", "统计"):
            self._show_stats()
            return

        # AI模型切换（任意阶段可用）
        ai_switch = re.match(r'(?:AI|ai|Ai)\s*[：:]\s*([A-Za-z0-9._-]+)', user_input)
        if ai_switch:
            self._switch_model(ai_switch.group(1).strip())
            return

        progress = GlobalState.get_progress()
        has_data = GlobalState.has("PRODUCT_INFO")

        if not has_data:
            self._handle_info_collection(user_input)
            return

        if progress["base_completed"] < 3:
            self._handle_base_flow(user_input, progress)
            return

        self._handle_branch_flow(user_input, progress)

    # ============ 模型与参数 ============

    def _get_model(self):
        info = GlobalState.get("PRODUCT_INFO")
        if info and info.get("ai_model"):
            return info["ai_model"]
        return self.DEFAULT_MODEL

    def _get_limits(self):
        settings = GlobalState.get("SETTINGS")
        if settings:
            return (
                settings.get("competitor_limit", 100),
                settings.get("review_limit", 1000),
                settings.get("keyword_limit", 100)
            )
        return 100, 1000, 100

    # ============ 模型切换与统计 ============

    def _switch_model(self, model_name):
        info = GlobalState.get("PRODUCT_INFO")
        if not info:
            info = {"description": "", "brand": "", "images": [], "reference_links": [], "ai_model": ""}
        old = info.get("ai_model", "") or self.DEFAULT_MODEL
        info["ai_model"] = model_name
        GlobalState.store("PRODUCT_INFO", info)
        with poe.start_message() as msg:
            msg.write(f"## 🤖 AI模型已切换\n\n")
            msg.write(f"- 原模型：{old}\n")
            msg.write(f"- 新模型：**{model_name}**\n\n")
            msg.write("后续所有分析阶段将使用新模型。\n")

    def _show_stats(self):
        progress = GlobalState.get_progress()
        base_done = progress.get("base_completed", 0)
        current_branch = progress.get("current_branch")
        completed = progress.get("completed_branches", [])
        all_mode = progress.get("all_mode", False)
        branch_stages = progress.get("branch_stages", {})
        model = self._get_model()

        with poe.start_message() as msg:
            msg.write("## 📊 流程进度\n\n")
            msg.write(f"**当前AI模型：** {model}\n\n")

            # 基础阶段
            msg.write("### 基础分析\n")
            for i, name in enumerate(self.base_stages):
                if i < base_done:
                    msg.write(f"- ✅ 阶段{i+1}：{name}\n")
                elif i == base_done:
                    msg.write(f"- 🔄 阶段{i+1}：{name}（进行中）\n")
                else:
                    msg.write(f"- ⬜ 阶段{i+1}：{name}\n")

            # 分支
            if base_done >= 3:
                msg.write("\n### 内容生成\n")
                branches = {"A": ("主图+附图Brief", self.branch_a_stages),
                            "B": ("A+页面Brief", self.branch_b_stages),
                            "C": ("标题+五点", self.branch_c_stages)}
                for bk, (label, stages) in branches.items():
                    done_count = branch_stages.get(bk, 0)
                    total = len(stages)
                    if bk in completed:
                        msg.write(f"- ✅ {label}（{total}/{total}）\n")
                    elif bk == current_branch:
                        msg.write(f"- 🔄 {label}（{done_count}/{total}）")
                        if done_count < total:
                            msg.write(f" — 下一步：{stages[done_count]}")
                        msg.write("\n")
                    else:
                        msg.write(f"- ⬜ {label}\n")

                if all_mode:
                    msg.write("\n📌 当前模式：**全部生成**\n")

            # GlobalState 状态概览
            msg.write("\n### 数据状态\n")
            state_keys = [
                ("PRODUCT_INFO", "产品信息"), ("REVIEW_EXTRACTION", "评论提取"),
                ("COMPETITOR_EXTRACTION", "竞品提取"), ("REFERENCE_ANALYSIS", "参考产品"),
                ("KEYWORDS", "关键词"), ("CONSUMER_INSIGHTS", "消费者洞察"),
                ("USAGE_FLOW", "使用流程"), ("SELLING_POINTS", "卖点列表"),
                ("IMAGE_PLAN", "图片规划"), ("IMAGE_DESIGN", "表达设计"),
                ("APLUS_MODULES", "A+模块"), ("KEYWORD_STRATEGY", "关键词策略"),
                ("TITLE", "标题方案"),
            ]
            for key, label in state_keys:
                data = GlobalState.get(key)
                if data:
                    if "_raw_text" in data:
                        msg.write(f"- ⚠️ {label}（降级文本）\n")
                    else:
                        msg.write(f"- ✅ {label}\n")

    # ============ AI 调用（核心：不传对话历史）============

    def _call_ai(self, prompt, system_msg=None, use_web_search=False):
        """调用主模型。不传对话历史，只传构建好的 prompt。
        输出实时流式显示给用户。返回完整文本。
        """
        messages = []
        if system_msg:
            messages.append(poe.Message(text=system_msg, sender="system"))
        if use_web_search:
            messages.append(poe.Message(text=prompt, parameters={"web_search": True}))
        else:
            messages.append(prompt)

        response = poe.call(self._get_model(), *messages, output=poe.default_chat)
        return response.text

    def _extract_conclusion(self, ai_output, extraction_prompt, required_fields=None):
        """用小模型从 AI 输出中提取结构化 JSON。"""
        full_prompt = f"""{extraction_prompt}

---

以下是需要提取的原始分析内容：

{ai_output}"""

        response = poe.call(self.EXTRACTION_MODEL, full_prompt)
        return parse_json_from_response(response.text, required_fields)

    # ============ 参数设置 ============

    def _handle_parameters(self, user_input):
        settings = {"competitor_limit": 100, "review_limit": 1000, "keyword_limit": 100}
        m = re.search(r'评论\s*[:：]?\s*(\d+)', user_input)
        if m:
            settings["review_limit"] = int(m.group(1))
        m = re.search(r'竞品\s*[:：]?\s*(\d+)', user_input)
        if m:
            settings["competitor_limit"] = int(m.group(1))
        m = re.search(r'关键词\s*[:：]?\s*(\d+)', user_input)
        if m:
            settings["keyword_limit"] = int(m.group(1))
        GlobalState.store("SETTINGS", settings)
        with poe.start_message() as msg:
            msg.write("## ⚙️ 参数设置成功\n\n")
            msg.write(f"📊 竞品数据：最多 {settings['competitor_limit']} 个\n")
            msg.write(f"💬 评论数据：最多 {settings['review_limit']} 条\n")
            msg.write(f"🔑 关键词数据：最多 {settings['keyword_limit']} 个\n\n")
            msg.write("请上传数据文件开始分析。\n")

    # ============ 信息收集 ============

    def _handle_info_collection(self, user_input):
        competitor_limit, review_limit, keyword_limit = self._get_limits()

        # --- Python 解析文件 ---
        competitor_data, review_data, keyword_data = [], [], []
        reference_links, product_images = [], []
        product_description, brand_name, ai_model = "", "", ""

        urls = re.findall(r'https?://[^\s]+', user_input)
        reference_links.extend(urls)
        remaining = re.sub(r'https?://[^\s]+', '', user_input).strip()
        if remaining:
            product_description = remaining

        # 提取品牌名
        if remaining:
            for pat in [
                r'(?:品牌名[称]?|品牌)\s*[：:是为]\s*(.+?)(?:\n|$|[,，。.、;；])',
                r'(?:brand\s*name|brand)\s*[：:=]\s*(.+?)(?:\n|$|[,，。.、;；])',
            ]:
                bm = re.search(pat, remaining, re.IGNORECASE)
                if bm:
                    brand_name = bm.group(1).strip()
                    break

        # 提取AI模型
        if remaining:
            am = re.search(r'(?:AI|ai|Ai)\s*[：:]\s*([A-Za-z0-9._-]+)', remaining)
            if am:
                ai_model = am.group(1).strip()

        # 解析附件
        for att in poe.query.attachments:
            fname = att.name.lower()

            if fname.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                product_images.append({'name': att.name, 'url': att.url})
                continue

            if fname.endswith(('.xlsx', '.xls')):
                try:
                    wb = openpyxl.load_workbook(BytesIO(att.get_contents()), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = rows[0]
                        hs = str(headers)
                        hu = hs.upper()
                        if '关键词' in hs and ('搜索量' in hs or '流量' in hs):
                            for row in rows[1:]:
                                if row and any(row):
                                    keyword_data.append(dict(zip(headers, row)))
                        elif 'ASIN' in hu and '评分数' in hs:
                            for row in rows[1:competitor_limit+1]:
                                if row and any(row):
                                    competitor_data.append(dict(zip(headers, row)))
                        elif 'ASIN' in hu and ('评论' in hs or '内容' in hs or '星级' in hs):
                            for row in rows[1:review_limit+1]:
                                if row and any(row):
                                    review_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取 {att.name} 出错：{str(e)[:100]}\n")

            elif fname.endswith('.csv'):
                try:
                    contents = att.get_contents()
                    try:
                        text = contents.decode('utf-8-sig')
                    except Exception:
                        text = contents.decode('gbk')
                    csv_rows = list(csv.DictReader(StringIO(text)))
                    if csv_rows:
                        h = list(csv_rows[0].keys())
                        hu = [x.upper() for x in h]
                        has_asin = 'ASIN' in hu
                        if has_asin and any(x in h for x in ['Rating', 'Body', 'Title', 'ReviewID']):
                            for row in csv_rows[:review_limit]:
                                review_data.append({
                                    'ASIN': row.get('ASIN', row.get('asin', '')),
                                    '标题': row.get('Title', row.get('标题', '')),
                                    '内容': row.get('Body', row.get('内容', '')),
                                    '星级': row.get('Rating', row.get('星级', '')),
                                    '评论时间': row.get('Date', row.get('评论时间', ''))
                                })
                        elif has_asin and any(x in h for x in ['商品标题', '评分数', '月销量']):
                            for row in csv_rows[:competitor_limit]:
                                competitor_data.append(row)
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取 {att.name} 出错：{str(e)[:100]}\n")

        if not (competitor_data or review_data or reference_links or product_images or product_description or keyword_data):
            with poe.start_message() as msg:
                msg.write("💡 请上传数据文件或粘贴链接。\n")
            return

        # --- 显示收集结果 ---
        with poe.start_message() as msg:
            msg.write("## ✅ 数据收集完成\n\n")
            msg.write(f"📊 竞品数据：{len(competitor_data)} 个\n")
            msg.write(f"💬 评论数据：{len(review_data)} 条\n")
            msg.write(f"🔗 参考链接：{len(reference_links)} 个\n")
            msg.write(f"🖼️ 产品图片：{len(product_images)} 张\n")
            msg.write(f"📝 产品描述：{'有' if product_description else '无'}\n")
            msg.write(f"🔑 关键词：{len(keyword_data)} 个\n")
            if brand_name:
                msg.write(f"✅ 品牌：{brand_name}\n")
            if ai_model:
                msg.write(f"🤖 AI模型：{ai_model}\n")
            msg.write("\n⏳ 正在进行AI结构化提取...\n")

        # --- 存储产品信息 ---
        GlobalState.store("PRODUCT_INFO", {
            "description": product_description,
            "brand": brand_name,
            "images": product_images,
            "reference_links": reference_links,
            "ai_model": ai_model
        })

        # --- AI 结构化提取（并行） ---
        tasks = []
        if review_data:
            tasks.append(lambda rd=review_data: self._extract_reviews(rd))
        if competitor_data:
            tasks.append(lambda cd=competitor_data: self._extract_competitors(cd))
        if reference_links:
            tasks.append(lambda rl=reference_links: self._analyze_references(rl))
        if keyword_data:
            tasks.append(lambda kd=keyword_data, kl=keyword_limit: self._process_keywords(kd, kl))
        if tasks:
            poe.parallel(*tasks, skip_exceptions=True)

        with poe.start_message() as msg:
            msg.write("---\n\n✅ 数据预处理完成！自动开始基础分析...\n\n")

        self._execute_base_stage(0)

    # ============ AI 结构化提取 ============

    def _extract_reviews(self, review_data):
        """11 维度结构化提取评论数据。"""
        review_text = "\n".join([
            f"[{r.get('星级', '?')}星] {r.get('标题', '')} {r.get('内容', '')}"
            for r in review_data
        ])
        if len(review_text) > 50000:
            review_text = review_text[:50000] + "\n...[已截断]"

        prompt = f"""你是一位专业的亚马逊评论数据分析师。请对以下 {len(review_data)} 条评论进行 11 维度结构化提取。

**重要：每个维度都必须认真填写，不能跳过或留空。**

{review_text}

---

请按以下 11 个维度逐一提取（使用中文分析，保留英文原话）：

## 1. 功能优点（Top 10）
按提及频率排序，每个附带英文原话示例和估计提及次数

## 2. 功能缺点（Top 10）
按提及频率排序，每个附带英文原话示例和估计提及次数

## 3. 使用场景
列出所有出现的使用场景及频率

## 4. 外观/设计评价
颜色、材质、外形、做工

## 5. 尺寸/便携性
大小、重量、收纳、兼容性（如杯架适配）

## 6. 价格感知
性价比评价

## 7. 与竞品对比
买家提到的品牌/产品对比

## 8. VOC 金句
真实英文原话：正面15条+负面10条+场景短语5条

## 9. 小众需求/意外用途
低频但有价值的特殊用途和需求

## 10. 复购/推荐意愿
买家忠诚度信号

## 11. 安装与维护逻辑
安装步骤、维护保养相关反馈"""

        with poe.start_message() as msg:
            msg.write("### 📝 评论数据结构化提取\n\n")

        ai_output = self._call_ai(prompt)

        extraction_prompt = """从以下评论分析中提取结构化 JSON。所有 11 个字段都必须填写。

## 示例输出（仅供格式参考，内容需来自实际分析）

```json
{
    "positive_features": [
        {"feature": "密封性好", "count": 180, "en_quotes": ["no leaks at all", "completely leak-proof even upside down"]},
        {"feature": "手感舒适", "count": 95, "en_quotes": ["fits perfectly in my hand", "comfortable grip"]}
    ],
    "negative_features": [
        {"feature": "盖子易损", "count": 60, "en_quotes": ["lid broke after 2 months", "cap cracked easily"]}
    ],
    "use_scenarios": [
        {"scenario": "办公室", "count": 120, "en_phrases": ["at my desk", "for the office"]}
    ],
    "appearance_feedback": ["外观简约大方", "颜色与图片一致"],
    "size_portability": ["适合标准杯架", "容量刚好"],
    "price_perception": "多数认为物有所值，部分觉得偏贵",
    "competitor_mentions": ["比X品牌更耐用", "不如Y品牌轻便"],
    "voc_gold_quotes": {
        "positive": ["best bottle I've ever owned", "keeps ice for 2 days"],
        "negative": ["paint started chipping", "dents too easily"],
        "scenario": ["perfect for hiking", "my gym essential"]
    },
    "niche_needs": ["有人用来装婴儿奶", "适合做蛋白粉摇杯"],
    "repurchase_signals": "约25%提到会回购或送人",
    "installation_maintenance": ["首次使用建议消毒", "不可放入洗碗机"]
}
```

## 格式要求
严格按示例的字段名和数据类型输出。count 为整数，en_quotes/en_phrases 为英文原话数组。

Output in ```json format:"""

        try:
            data = self._extract_conclusion(ai_output, extraction_prompt,
                required_fields=["positive_features", "negative_features", "voc_gold_quotes"])
            GlobalState.store("REVIEW_EXTRACTION", data)
            with poe.start_message() as msg:
                msg.write("✅ 评论提取完成\n")
        except Exception as e:
            GlobalState.store("REVIEW_EXTRACTION", {"_raw_text": ai_output[:8000]})
            with poe.start_message() as msg:
                msg.write(f"⚠️ 评论JSON提取降级: {str(e)[:80]}\n")

    def _extract_competitors(self, competitor_data):
        """结构化提取竞品数据。"""
        comp_text = "\n".join([
            f"{i+1}. 标题：{c.get('商品标题', '')}\n   卖点：{c.get('产品卖点', '')}"
            for i, c in enumerate(competitor_data)
        ])
        if len(comp_text) > 30000:
            comp_text = comp_text[:30000] + "\n...[已截断]"

        prompt = f"""你是一位亚马逊竞品分析专家。请对以下 {len(competitor_data)} 个竞品进行结构化分析：

{comp_text}

---

请输出（使用中文）：

## 1. 产品类型与市场定位

## 2. 高频卖点统计（Top 15）
按出现频率排序

## 3. 设计趋势
材质、功能、风格、价格区间

## 4. 功能维度对比
按维度列出各竞品表现

## 5. 代表性竞品（3-5个）

## 6. 市场空白与机会"""

        with poe.start_message() as msg:
            msg.write("### 🏆 竞品数据结构化提取\n\n")

        ai_output = self._call_ai(prompt)

        extraction_prompt = """从以下竞品分析中提取结构化 JSON。

Output in ```json format:
```json
{
    "product_type": "产品类型总结",
    "top_selling_points": [{"point": "卖点", "frequency": "频率描述"}],
    "design_trends": {"materials": "材质", "functions": "功能", "styles": "风格", "price_range": "价格"},
    "feature_comparison": [{"dimension": "维度", "details": "对比详情"}],
    "representative_products": [{"title": "标题", "core_points": "核心卖点"}],
    "market_gaps": ["市场空白"]
}
```"""

        try:
            data = self._extract_conclusion(ai_output, extraction_prompt,
                required_fields=["top_selling_points", "market_gaps"])
            GlobalState.store("COMPETITOR_EXTRACTION", data)
            with poe.start_message() as msg:
                msg.write("✅ 竞品提取完成\n")
        except Exception as e:
            GlobalState.store("COMPETITOR_EXTRACTION", {"_raw_text": ai_output[:6000]})
            with poe.start_message() as msg:
                msg.write(f"⚠️ 竞品JSON提取降级: {str(e)[:80]}\n")

    def _analyze_references(self, reference_links):
        """使用 web_search 分析参考产品。"""
        links_text = "\n".join([f"- {link}" for link in reference_links])

        prompt = f"""请访问以下亚马逊产品链接，获取并分析产品信息：

{links_text}

对于每个产品，请提取：

## 参考产品：[产品名称]
### 基本信息
完整标题、价格、评分、评论数

### 五点描述（Bullet Points）
原文逐条列出

### 核心卖点分析
主打卖点、差异化特点、目标人群

### 主图分析
主图表达、视觉风格

---
## 参考产品总结
共同特征、可借鉴之处、可改进之处

要求：务必访问链接获取真实信息。"""

        with poe.start_message() as msg:
            msg.write("### 🔗 参考产品分析\n\n")

        ai_output = None
        for attempt in range(3):
            try:
                ai_output = self._call_ai(prompt, use_web_search=True)
                break
            except Exception:
                if attempt < 2:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 重试中（{attempt+2}/3）...\n")
                    time.sleep(2)

        if not ai_output:
            with poe.start_message() as msg:
                msg.write("⚠️ 参考产品分析失败\n")
            return

        extraction_prompt = """从以下参考产品分析中提取结构化 JSON。

Output in ```json format:
```json
{
    "products": [
        {
            "title": "完整标题",
            "price": "价格",
            "rating": "评分",
            "review_count": "评论数",
            "bullet_points": ["五点1", "五点2", "五点3", "五点4", "五点5"],
            "main_selling_point": "主打卖点",
            "differentiator": "差异化",
            "target_audience": "目标人群",
            "visual_style": "视觉风格"
        }
    ],
    "common_features": "共同特征",
    "learnable_aspects": "可借鉴",
    "improvable_aspects": "可改进"
}
```"""

        try:
            data = self._extract_conclusion(ai_output, extraction_prompt, required_fields=["products"])
            GlobalState.store("REFERENCE_ANALYSIS", data)
            with poe.start_message() as msg:
                msg.write("✅ 参考产品分析完成\n")
        except Exception as e:
            GlobalState.store("REFERENCE_ANALYSIS", {"_raw_text": ai_output[:8000]})
            with poe.start_message() as msg:
                msg.write(f"⚠️ 参考产品JSON提取降级: {str(e)[:80]}\n")

    def _process_keywords(self, keyword_data, keyword_limit):
        """Python 处理关键词（不需要AI）。"""
        filtered = [kw for kw in keyword_data if kw.get('相关性档位', '') in ['强相关', '高相关']]

        def vol(kw):
            v = kw.get('周搜索量', 0)
            return int(v) if v else 0

        filtered.sort(key=vol, reverse=True)
        final = filtered[:keyword_limit]

        kw_json = []
        for kw in final:
            name = kw.get('关键词 (数据来源于西柚找词)', kw.get('关键词', ''))
            kw_json.append({
                "keyword": name,
                "translation": kw.get('翻译', ''),
                "relevance": kw.get('相关性档位', ''),
                "volume": vol(kw)
            })

        GlobalState.store("KEYWORDS", {
            "keywords": kw_json,
            "stats": {"original": len(keyword_data), "filtered": len(filtered), "final": len(final)}
        })

        with poe.start_message() as msg:
            msg.write(f"### 🔑 关键词数据\n\n")
            msg.write(f"原始 {len(keyword_data)} → 筛选 {len(filtered)} → 保留 {len(final)}\n\n")
            msg.write("| # | 关键词 | 翻译 | 相关性 | 周搜索量 |\n")
            msg.write("|---|--------|------|--------|----------|\n")
            for i, kw in enumerate(kw_json[:20], 1):
                msg.write(f"| {i} | {kw['keyword']} | {kw['translation']} | {kw['relevance']} | {kw['volume']} |\n")
            if len(kw_json) > 20:
                msg.write(f"\n... 共 {len(kw_json)} 个\n")
            msg.write("\n✅ 关键词处理完成\n")

    # ============ 基础阶段 ============

    def _handle_base_flow(self, user_input, progress):
        current = progress["base_completed"]

        if "继续" in user_input or "下一步" in user_input:
            if current >= 3:
                self._show_branch_menu()
            else:
                self._execute_base_stage(current)
        elif "重新生成" in user_input:
            stage_to_redo = max(current - 1, 0)
            # 提取用户附带的修改意见（去掉"重新生成"关键词）
            feedback = re.sub(r'重新生成[，,。.：:\s]*', '', user_input).strip()
            if not feedback:
                feedback = None
            self._execute_base_stage(stage_to_redo, feedback=feedback)
        else:
            self._handle_discussion(user_input, progress)

    def _execute_base_stage(self, stage_num, feedback=None):
        # 校验前置数据
        required = ["PRODUCT_INFO"]
        if stage_num >= 1:
            required.append("CONSUMER_INSIGHTS")
        if stage_num >= 2:
            required.append("USAGE_FLOW")
        GlobalState.require(*required)

        stage_name = self.base_stages[stage_num]

        with poe.start_message() as msg:
            if feedback:
                msg.write(f"# 🔄 重新生成：基础阶段 {stage_num+1} - {stage_name}\n\n")
                msg.write(f"📝 修改意见：{feedback}\n\n")
            else:
                msg.write(f"# ⚙️ 基础阶段 {stage_num+1}：{stage_name}\n\n")

        # 检查前置数据降级情况
        degraded = GlobalState.check_degraded("REVIEW_EXTRACTION", "COMPETITOR_EXTRACTION")
        if degraded:
            labels = {"REVIEW_EXTRACTION": "评论提取", "COMPETITOR_EXTRACTION": "竞品提取"}
            names = [labels[k] for k in degraded]
            with poe.start_message() as msg:
                msg.write(f"⚠️ **注意：以下数据为降级模式（原始文本而非结构化JSON），可能影响分析质量：**{', '.join(names)}\n\n")

        prompt = self._build_base_prompt(stage_num, feedback)

        t0 = time.time()
        ai_output = self._call_ai(prompt)
        elapsed = time.time() - t0

        # 用小模型提取结论JSON
        key, ext_prompt, req_fields = self._get_base_extraction_config(stage_num)

        extraction_ok = True
        try:
            conclusion = self._extract_conclusion(ai_output, ext_prompt, req_fields)
            GlobalState.store(key, conclusion)
        except Exception:
            GlobalState.store(key, {"_raw_text": ai_output[:6000]})
            extraction_ok = False

        # 更新进度
        progress = GlobalState.get_progress()
        progress["base_completed"] = stage_num + 1
        GlobalState.set_progress(progress)

        with poe.start_message() as msg:
            msg.write(f"\n---\n\n")
            msg.write(f"✅ **基础阶段 {stage_num+1} 完成**：{stage_name}\n")
            msg.write(f"⏱️ {elapsed:.1f}s | 输入 {len(prompt):,} 字符 | 输出 {len(ai_output):,} 字符\n")
            if not extraction_ok:
                msg.write("⚠️ JSON提取降级（不影响使用）\n")
            msg.write("\n")
            if stage_num < 2:
                msg.write("💡 输入讨论内容，或「**重新生成**」重做，或「**继续**」进入下一阶段。\n")
            else:
                msg.write("💡 输入讨论内容，或「**重新生成**」重做，或「**继续**」选择生成内容。\n")

    def _build_base_prompt(self, stage_num, feedback=None):
        """从 GlobalState 精确读取所需字段，构建 prompt。"""
        product_info = GlobalState.get("PRODUCT_INFO") or {}
        review = GlobalState.get("REVIEW_EXTRACTION") or {}
        competitor = GlobalState.get("COMPETITOR_EXTRACTION") or {}
        reference = GlobalState.get("REFERENCE_ANALYSIS") or {}

        product_section = self._fmt_product_info(product_info)

        if stage_num == 0:
            prompt = f"""{product_section}

## 评论数据（11维度提取结果）
{self._fmt_review(review)}

## 竞品数据提取结果
{self._fmt_competitor(competitor)}

## 参考产品分析
{self._fmt_reference(reference)}

---

# 🎯 任务：消费者洞察分析

你是资深亚马逊运营专家。请基于上述数据进行深度消费者洞察分析。

## 输出要求（中文）

## 1. 消费者画像
- 主要用户群体特征、使用场景、购买动机

## 2. 购物决策因素
- 最重要的3-5个决策因素及原因

## 3. 核心痛点
- 按优先级排序，附评论英文原话

## 4. 期望与需求
- 明确需求和隐含期望

要求：基于数据，直白表达，避免主观臆测。"""

        elif stage_num == 1:
            consumer = GlobalState.get("CONSUMER_INSIGHTS") or {}
            prompt = f"""{product_section}

## 消费者洞察结论
{self._fmt_json(consumer)}

## 参考产品分析
{self._fmt_reference(reference)}

---

# 🎯 任务：使用流程复现

你是产品体验设计专家。请完整复现产品使用流程。

## 输出要求（中文）

## 使用前
开箱、安装、准备工作

## 使用中
具体使用步骤，每步操作要点

## 使用后
收纳、维护、清洁保养

要求：不遗漏步骤，代入消费者视角，具体可操作。"""

        elif stage_num == 2:
            consumer = GlobalState.get("CONSUMER_INSIGHTS") or {}
            usage = GlobalState.get("USAGE_FLOW") or {}
            prompt = f"""{product_section}

## 使用流程结论
{self._fmt_json(usage)}

## 消费者洞察结论
{self._fmt_json(consumer)}

## 竞品高频卖点
{self._fmt_competitor_points(competitor)}

## 参考产品分析
{self._fmt_reference(reference)}

---

# 🎯 任务：卖点提取

你是产品竞争分析专家。请基于使用流程每个步骤提取卖点。

## 输出要求（中文）

### 第一部分：基于使用流程的卖点
对每个使用步骤分析：
1. 涉及的产品功能
2. 与竞品对比优势
3. 可表达的卖点

### 第二部分：使用流程之外的卖点
⚠️ 以下维度的卖点不一定出现在使用步骤中，但对购买决策同样重要，请单独检查：
- **品牌与信任**：品牌故事、认证、质保承诺
- **包装与开箱体验**：礼盒包装、配件赠品、第一印象
- **售后服务**：退换政策、客服响应、保修期
- **社会认同**：销量、评分、网红推荐、获奖
- **环保/健康属性**：材质安全、可回收、BPA-free 等

### 最终输出

## 卖点汇总（按重要性排序，合并两部分）
1. [卖点名] — 证据/理由
2. ...

要求：具体可量化，对比有依据，避免空洞形容词。"""
        else:
            prompt = "[未定义]"

        if feedback:
            prompt += f"\n\n---\n⚠️ **用户反馈（重新生成）：** {feedback}\n\n请在分析中重点关注和改进以上反馈提到的内容，同时保持其他部分的完整性。"
        return prompt

    def _get_base_extraction_config(self, stage_num):
        """返回 (state_key, extraction_prompt, required_fields)。"""
        if stage_num == 0:
            return "CONSUMER_INSIGHTS", """提取消费者洞察的结构化结论。

Output in ```json format:
```json
{
    "personas": [{"group": "用户群体", "characteristics": "特征", "scenarios": ["场景"], "motivation": "动机"}],
    "decision_factors": [{"factor": "因素", "importance": "重要程度", "reason": "原因"}],
    "pain_points": [{"pain": "痛点", "priority": 1, "manifestation": "表现", "en_quotes": ["英文原话"]}],
    "expectations": {"explicit": ["明确需求"], "implicit": ["隐含期望"]}
}
```""", ["personas", "decision_factors", "pain_points"]

        elif stage_num == 1:
            return "USAGE_FLOW", """提取使用流程的结构化结论。

Output in ```json format:
```json
{
    "pre_use": [{"step": "步骤", "details": "说明", "notes": "注意事项"}],
    "during_use": [{"step": "步骤", "details": "说明", "key_points": "要点"}],
    "post_use": [{"step": "步骤", "details": "说明", "frequency": "频率"}]
}
```""", ["pre_use", "during_use", "post_use"]

        else:
            return "SELLING_POINTS", """提取所有卖点，不要遗漏。按重要性排序。
注意：usage_step 字段允许为"N/A"——部分卖点（如品牌信任、包装体验、售后保障等）不直接对应使用步骤，也需要提取。

## 示例输出（仅供格式参考）

```json
{
    "selling_points": [
        {
            "rank": 1,
            "name": "72小时超长保温",
            "description": "远超行业标准的保温时长",
            "evidence": "评论中200+买家提及保温效果，竞品普遍仅12-24小时",
            "usage_step": "使用中-饮用",
            "target_audience": "户外运动爱好者、长途通勤族",
            "vs_competitors": "竞品最高24小时，我们是3倍",
            "keywords_en": ["72-hour insulation", "triple insulated"]
        },
        {
            "rank": 5,
            "name": "品牌售后保障",
            "description": "终身质保承诺",
            "evidence": "竞品多为1年保修，终身质保形成信任差异",
            "usage_step": "N/A",
            "target_audience": "注重长期价值的消费者",
            "vs_competitors": "竞品普遍1年有限保修",
            "keywords_en": ["lifetime warranty", "guaranteed"]
        }
    ]
}
```

## 格式要求
严格按示例的字段名和数据类型。rank 为整数，keywords_en 为英文数组。

Output in ```json format:""", ["selling_points"]

    # ============ 数据格式化 ============

    def _fmt_product_info(self, info):
        parts = ["# 产品基本信息\n"]
        if info.get("brand"):
            parts.append(f"**⚠️ 品牌名称：{info['brand']}**（所有内容必须使用此品牌名）\n")
        if info.get("description"):
            parts.append(f"**产品描述**：{info['description']}\n")
        if info.get("images"):
            parts.append(f"**产品图片**：{len(info['images'])} 张")
            for img in info["images"]:
                parts.append(f"  - {img['name']}: {img['url']}")
        return "\n".join(parts)

    def _fmt_review(self, review):
        if not review:
            return "[无评论数据]"
        if "_raw_text" in review:
            return review["_raw_text"]

        parts = []
        for key, label in [("positive_features", "功能优点"), ("negative_features", "功能缺点")]:
            items = review.get(key, [])
            if items:
                parts.append(f"\n**{label}：**")
                for f in items[:10]:
                    qs = ', '.join(f.get('en_quotes', [])[:2])
                    parts.append(f"- {f.get('feature', '?')}（~{f.get('count', '?')}次）{qs}")

        for key, label in [("use_scenarios", "使用场景")]:
            items = review.get(key, [])
            if items:
                parts.append(f"\n**{label}：**")
                for s in items:
                    parts.append(f"- {s.get('scenario', '?')}（~{s.get('count', '?')}次）")

        simple_fields = [
            ("appearance_feedback", "外观/设计"), ("size_portability", "尺寸/便携性"),
            ("competitor_mentions", "竞品对比"), ("niche_needs", "小众需求/意外用途"),
            ("installation_maintenance", "安装与维护")
        ]
        for key, label in simple_fields:
            items = review.get(key, [])
            if items:
                parts.append(f"\n**{label}：** {'; '.join(str(x) for x in items[:5])}")

        if review.get("price_perception"):
            parts.append(f"\n**价格感知：** {review['price_perception']}")
        if review.get("repurchase_signals"):
            parts.append(f"\n**复购意愿：** {review['repurchase_signals']}")

        voc = review.get("voc_gold_quotes", {})
        if voc:
            if voc.get("positive"):
                parts.append(f"\n**VOC正面：** {', '.join(str(x) for x in voc['positive'][:10])}")
            if voc.get("negative"):
                parts.append(f"\n**VOC负面：** {', '.join(str(x) for x in voc['negative'][:8])}")
            if voc.get("scenario"):
                parts.append(f"\n**VOC场景：** {', '.join(str(x) for x in voc['scenario'][:5])}")

        return "\n".join(parts) if parts else "[评论数据为空]"

    def _fmt_competitor(self, comp):
        if not comp:
            return "[无竞品数据]"
        if "_raw_text" in comp:
            return comp["_raw_text"]

        parts = []
        if comp.get("product_type"):
            parts.append(f"**产品类型：** {comp['product_type']}")
        if comp.get("top_selling_points"):
            parts.append("\n**高频卖点：**")
            for sp in comp["top_selling_points"][:10]:
                parts.append(f"- {sp.get('point', '?')}（{sp.get('frequency', '')}）")
        t = comp.get("design_trends", {})
        if t:
            parts.append(f"\n**趋势：** 材质={t.get('materials','?')} | 功能={t.get('functions','?')} | 价格={t.get('price_range','?')}")
        if comp.get("market_gaps"):
            parts.append(f"\n**市场空白：** {'; '.join(str(x) for x in comp['market_gaps'][:5])}")
        return "\n".join(parts) if parts else "[竞品数据为空]"

    def _fmt_competitor_points(self, comp):
        if not comp:
            return "[无]"
        if "_raw_text" in comp:
            return comp["_raw_text"][:2000]
        lines = []
        if comp.get("top_selling_points"):
            for sp in comp["top_selling_points"][:10]:
                lines.append(f"- {sp.get('point', '?')}（{sp.get('frequency', '')}）")
        if comp.get("market_gaps"):
            lines.append(f"\n**市场空白：** {'; '.join(str(x) for x in comp['market_gaps'][:5])}")
        return "\n".join(lines) if lines else "[无]"

    def _fmt_reference(self, ref):
        if not ref:
            return "[无参考产品]"
        if "_raw_text" in ref:
            return ref["_raw_text"]

        parts = []
        for p in ref.get("products", []):
            parts.append(f"### {p.get('title', '?')}")
            parts.append(f"价格：{p.get('price','?')} | 评分：{p.get('rating','?')} | 评论：{p.get('review_count','?')}")
            bps = p.get("bullet_points", [])
            if bps:
                parts.append("**五点：**")
                for i, bp in enumerate(bps, 1):
                    parts.append(f"  {i}. {bp}")
            parts.append(f"**主打：** {p.get('main_selling_point','?')}")
            parts.append(f"**差异化：** {p.get('differentiator','?')}")
            parts.append(f"**人群：** {p.get('target_audience','?')}")
            parts.append(f"**视觉：** {p.get('visual_style','?')}\n")

        if ref.get("learnable_aspects"):
            parts.append(f"**可借鉴：** {ref['learnable_aspects']}")
        if ref.get("improvable_aspects"):
            parts.append(f"**可改进：** {ref['improvable_aspects']}")
        return "\n".join(parts) if parts else "[无]"

    def _fmt_json(self, data):
        if not data:
            return "[无]"
        if "_raw_text" in data:
            return data["_raw_text"]
        return json.dumps(data, ensure_ascii=False, indent=2)

    # ============ 讨论模式（阶段隔离）============

    def _handle_discussion(self, user_input, progress):
        """阶段隔离讨论：只传当前阶段结论 + 必要上下文，不传对话历史。

        讨论输入构成：
        - 当前阶段名称和结论（JSON）
        - 前置数据的摘要（只列维度标题，不展开）
        - 用户问题
        """
        base_done = progress.get("base_completed", 0)
        current_branch = progress.get("current_branch")

        # 确定当前所在阶段和对应结论
        stage_label, conclusion, data_summary = self._get_discussion_context(progress)

        system_msg = """你是资深亚马逊运营专家。用户正在与你讨论某个分析阶段的结论。
规则：
1. 只基于提供的结论数据回答，不要编造数据
2. 如果用户提出修改建议，说明采纳后会如何影响结论
3. 如果用户的问题超出当前阶段范围，提示可以在对应阶段讨论
4. 保持简洁，直接回答"""

        prompt = f"""## 当前阶段：{stage_label}

## 阶段结论
{self._fmt_json(conclusion)}

{data_summary}

---

**用户问题/意见：**
{user_input}

请回答。如果用户提出了具体修改意见，请说明建议是否合理，以及用户可以「重新生成 + 修改意见」来应用修改。"""

        self._call_ai(prompt, system_msg=system_msg)

        with poe.start_message() as msg:
            msg.write("\n---\n")
            if current_branch:
                msg.write("💡 继续讨论，或「**重新生成**」（可附修改意见），或「**继续**」下一阶段。\n")
            else:
                next_label = "选择生成内容" if base_done >= 3 else "下一阶段"
                msg.write(f"💡 继续讨论，或「**重新生成**」（可附修改意见），或「**继续**」{next_label}。\n")

    def _get_discussion_context(self, progress):
        """根据进度确定讨论上下文：(stage_label, conclusion_data, data_summary_str)。"""
        base_done = progress.get("base_completed", 0)
        current_branch = progress.get("current_branch")
        branch_stages_map = progress.get("branch_stages", {})

        # 分支讨论
        if current_branch:
            stages = self._get_branch_stages(current_branch)
            stage_idx = max(branch_stages_map.get(current_branch, 1) - 1, 0)
            stage_label = f"分支{current_branch} - {stages[stage_idx]}"

            keys_map = {
                "A": ["IMAGE_PLAN", "IMAGE_DESIGN", None],
                "B": ["APLUS_MODULES", None],
                "C": ["KEYWORD_STRATEGY", "TITLE", None]
            }
            branch_keys = keys_map.get(current_branch, [])
            conclusion = None
            if stage_idx < len(branch_keys) and branch_keys[stage_idx]:
                conclusion = GlobalState.get(branch_keys[stage_idx])

            data_summary = self._build_data_summary(include_base=True)
            return stage_label, conclusion, data_summary

        # 基础阶段讨论
        if base_done > 0:
            last_stage = base_done - 1
            stage_label = f"基础阶段{last_stage+1} - {self.base_stages[last_stage]}"
            keys = ["CONSUMER_INSIGHTS", "USAGE_FLOW", "SELLING_POINTS"]
            conclusion = GlobalState.get(keys[last_stage])

            # 提供前置阶段的简短摘要（不展开JSON）
            data_summary = self._build_data_summary(include_base=False, up_to_stage=last_stage)
            return stage_label, conclusion, data_summary

        return "信息收集", None, ""

    def _build_data_summary(self, include_base=False, up_to_stage=None):
        """构建简要的数据摘要（只列有哪些数据，不展开内容，节省token）。"""
        parts = ["## 可用数据摘要"]

        # 原始数据
        review = GlobalState.get("REVIEW_EXTRACTION")
        if review and "_raw_text" not in review:
            pos = len(review.get("positive_features", []))
            neg = len(review.get("negative_features", []))
            parts.append(f"- 评论提取：{pos}个优点, {neg}个缺点")
        competitor = GlobalState.get("COMPETITOR_EXTRACTION")
        if competitor and "_raw_text" not in competitor:
            sp_count = len(competitor.get("top_selling_points", []))
            parts.append(f"- 竞品提取：{sp_count}个高频卖点")
        if GlobalState.has("REFERENCE_ANALYSIS"):
            parts.append("- 参考产品分析：有")
        if GlobalState.has("KEYWORDS"):
            kw = GlobalState.get("KEYWORDS")
            parts.append(f"- 关键词：{len(kw.get('keywords', []))}个")

        # 基础阶段结论摘要
        if include_base or up_to_stage is not None:
            base_keys = [
                ("CONSUMER_INSIGHTS", "消费者洞察"),
                ("USAGE_FLOW", "使用流程"),
                ("SELLING_POINTS", "卖点列表")
            ]
            limit = 3 if include_base else (up_to_stage if up_to_stage is not None else 0)
            for i, (key, label) in enumerate(base_keys):
                if i >= limit:
                    break
                data = GlobalState.get(key)
                if data:
                    parts.append(f"- {label}：已完成" + ("（降级）" if "_raw_text" in data else ""))

        return "\n".join(parts) if len(parts) > 1 else ""

    def _show_branch_menu(self):
        """分支菜单（第2步实现完整分支逻辑）。"""
        progress = GlobalState.get_progress()
        with poe.start_message() as msg:
            msg.write("## 🎯 基础分析完成！选择生成内容\n\n")
            msg.write("| 选项 | 内容 | 说明 |\n")
            msg.write("|------|------|------|\n")
            msg.write("| **1** | 主图+附图设计Brief | 7张图片设计方案 |\n")
            msg.write("| **2** | A+页面设计Brief | A+页面模块方案 |\n")
            msg.write("| **3** | 标题+五点描述 | Listing文案优化 |\n")
            msg.write("| **4** | 全部生成 | 按顺序生成所有 |\n\n")
            done = progress.get("completed_branches", [])
            if done:
                msg.write(f"✅ 已完成：{', '.join(done)}\n\n")
            if not GlobalState.has("KEYWORDS"):
                msg.write("⚠️ 生成「标题+五点」需先上传关键词数据\n\n")
            msg.write("请输入数字（1-4）：\n")

    def _handle_branch_flow(self, user_input, progress):
        """分支流程控制。"""
        current_branch = progress.get("current_branch")

        # 如果在某个分支中
        if current_branch:
            branch_stages = progress.get("branch_stages", {})
            current_stage = branch_stages.get(current_branch, 0)
            stages = self._get_branch_stages(current_branch)

            if "继续" in user_input or "下一步" in user_input:
                if current_stage >= len(stages):
                    self._complete_branch(current_branch, progress)
                else:
                    self._execute_branch_stage(current_branch, current_stage)
            elif "重新生成" in user_input:
                stage_to_redo = max(current_stage - 1, 0)
                feedback = re.sub(r'重新生成[，,。.：:\s]*', '', user_input).strip()
                if not feedback:
                    feedback = None
                self._execute_branch_stage(current_branch, stage_to_redo, feedback=feedback)
            else:
                self._handle_discussion(user_input, progress)
            return

        # 等待关键词上传
        if progress.get("waiting_keywords"):
            self._handle_keyword_upload(user_input, progress)
            return

        # 解析分支选择
        choice = self._parse_branch_choice(user_input)

        if choice == "A":
            self._start_branch("A", progress)
        elif choice == "B":
            self._start_branch("B", progress)
        elif choice == "C":
            if not GlobalState.has("KEYWORDS"):
                self._request_keyword_upload(progress)
            else:
                self._start_branch("C", progress)
        elif choice == "ALL":
            progress["all_mode"] = True
            GlobalState.set_progress(progress)
            self._start_branch("A", progress)
        else:
            self._show_branch_menu()

    def _parse_branch_choice(self, user_input):
        ui = user_input.lower()
        if "1" in ui or "主图" in ui or "附图" in ui:
            return "A"
        elif "2" in ui or "a+" in ui:
            return "B"
        elif "3" in ui or "标题" in ui or "五点" in ui:
            return "C"
        elif "4" in ui or "全部" in ui or "都要" in ui:
            return "ALL"
        return None

    def _get_branch_stages(self, branch):
        if branch == "A":
            return self.branch_a_stages
        elif branch == "B":
            return self.branch_b_stages
        return self.branch_c_stages

    def _start_branch(self, branch, progress):
        labels = {"A": "主图+附图设计Brief", "B": "A+页面设计Brief", "C": "标题+五点描述"}
        with poe.start_message() as msg:
            msg.write(f"## 🎨 开始生成：{labels[branch]}\n\n")
        progress["current_branch"] = branch
        bs = progress.get("branch_stages", {})
        bs[branch] = 0
        progress["branch_stages"] = bs
        GlobalState.set_progress(progress)
        self._execute_branch_stage(branch, 0)

    def _execute_branch_stage(self, branch, stage_num, feedback=None):
        # 校验前置数据：所有分支都需要基础阶段的产出
        GlobalState.require("PRODUCT_INFO", "SELLING_POINTS")
        if branch == "C":
            GlobalState.require("KEYWORDS")

        stages = self._get_branch_stages(branch)
        stage_name = stages[stage_num]

        with poe.start_message() as msg:
            if feedback:
                msg.write(f"# 🔄 重新生成：分支{branch}-阶段 {stage_num+1} - {stage_name}\n\n")
                msg.write(f"📝 修改意见：{feedback}\n\n")
            else:
                msg.write(f"# ⚙️ 分支{branch}-阶段 {stage_num+1}：{stage_name}\n\n")

        # 检查前置数据降级
        check_keys = ["SELLING_POINTS", "CONSUMER_INSIGHTS"]
        degraded = GlobalState.check_degraded(*check_keys)
        if degraded:
            labels = {"SELLING_POINTS": "卖点列表", "CONSUMER_INSIGHTS": "消费者洞察"}
            names = [labels.get(k, k) for k in degraded]
            with poe.start_message() as msg:
                msg.write(f"⚠️ **注意：以下前置数据为降级模式，可能影响生成质量：**{', '.join(names)}\n\n")

        prompt = self._build_branch_prompt(branch, stage_num, feedback)

        t0 = time.time()
        ai_output = self._call_ai(prompt)
        elapsed = time.time() - t0

        key, ext_prompt, req_fields = self._get_branch_extraction_config(branch, stage_num)

        extraction_ok = True
        if key:
            try:
                conclusion = self._extract_conclusion(ai_output, ext_prompt, req_fields)
                GlobalState.store(key, conclusion)
            except Exception:
                GlobalState.store(key, {"_raw_text": ai_output[:6000]})
                extraction_ok = False

        progress = GlobalState.get_progress()
        bs = progress.get("branch_stages", {})
        bs[branch] = stage_num + 1
        progress["branch_stages"] = bs
        GlobalState.set_progress(progress)

        with poe.start_message() as msg:
            msg.write(f"\n---\n\n")
            msg.write(f"✅ **分支{branch}-阶段 {stage_num+1} 完成**：{stage_name}\n")
            msg.write(f"⏱️ {elapsed:.1f}s | 输入 {len(prompt):,} 字符 | 输出 {len(ai_output):,} 字符\n")
            if key and not extraction_ok:
                msg.write("⚠️ JSON提取降级\n")
            msg.write("\n")
            if stage_num < len(stages) - 1:
                msg.write("💡 输入讨论内容，或「**重新生成**」重做，或「**继续**」下一阶段。\n")
            else:
                msg.write("💡 输入讨论内容，或「**重新生成**」重做，或「**继续**」完成此分支。\n")

    def _complete_branch(self, branch, progress):
        labels = {"A": "主图+附图设计Brief", "B": "A+页面设计Brief", "C": "标题+五点描述"}
        completed = progress.get("completed_branches", [])
        if branch not in completed:
            completed.append(branch)
        progress["completed_branches"] = completed
        progress["current_branch"] = None
        GlobalState.set_progress(progress)

        with poe.start_message() as msg:
            msg.write(f"✅ **{labels[branch]} 已完成！**\n\n")

        if progress.get("all_mode"):
            next_b = None
            if branch == "A" and "B" not in completed:
                next_b = "B"
            elif branch == "B" and "C" not in completed:
                if GlobalState.has("KEYWORDS"):
                    next_b = "C"
                else:
                    with poe.start_message() as msg:
                        msg.write("⚠️ 缺少关键词数据，无法自动生成标题五点。\n\n")
            if next_b:
                with poe.start_message() as msg:
                    msg.write("正在继续生成下一项...\n\n")
                self._start_branch(next_b, progress)
                return

        self._show_branch_menu()

    def _request_keyword_upload(self, progress):
        progress["waiting_keywords"] = True
        GlobalState.set_progress(progress)
        with poe.start_message() as msg:
            msg.write("## 🔑 需要关键词数据\n\n")
            msg.write("生成「标题+五点」需要关键词Excel（含 关键词、翻译、相关性档位、周搜索量）。\n")
            msg.write("请上传文件。\n")

    def _handle_keyword_upload(self, user_input, progress):
        keyword_limit = self._get_limits()[2]
        keyword_data = []
        for att in poe.query.attachments:
            if att.name.lower().endswith(('.xlsx', '.xls')):
                try:
                    wb = openpyxl.load_workbook(BytesIO(att.get_contents()), read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = rows[0]
                        for row in rows[1:]:
                            if row and any(row):
                                keyword_data.append(dict(zip(headers, row)))
                    wb.close()
                except Exception as e:
                    with poe.start_message() as msg:
                        msg.write(f"⚠️ 读取出错：{str(e)[:100]}\n")
        if keyword_data:
            self._process_keywords(keyword_data, keyword_limit)
            progress["waiting_keywords"] = False
            GlobalState.set_progress(progress)
            self._start_branch("C", progress)
        elif GlobalState.has("KEYWORDS"):
            progress["waiting_keywords"] = False
            GlobalState.set_progress(progress)
            self._start_branch("C", progress)
        else:
            with poe.start_message() as msg:
                msg.write("⚠️ 未检测到关键词数据，请上传Excel。\n")


    # ============ 分支 Prompt 构建 ============

    def _build_branch_prompt(self, branch, stage_num, feedback=None):
        if branch == "A":
            prompt = self._build_branch_a_prompt(stage_num)
        elif branch == "B":
            prompt = self._build_branch_b_prompt(stage_num)
        else:
            prompt = self._build_branch_c_prompt(stage_num)
        if feedback:
            prompt += f"\n\n---\n⚠️ **用户反馈（重新生成）：** {feedback}\n\n请在分析中重点关注和改进以上反馈提到的内容，同时保持其他部分的完整性。"
        return prompt

    def _build_branch_a_prompt(self, stage_num):
        product_info = GlobalState.get("PRODUCT_INFO") or {}
        selling_points = GlobalState.get("SELLING_POINTS") or {}
        consumer = GlobalState.get("CONSUMER_INSIGHTS") or {}
        reference = GlobalState.get("REFERENCE_ANALYSIS") or {}
        ps = self._fmt_product_info(product_info)

        if stage_num == 0:
            return f"""{ps}

## 卖点列表
{self._fmt_selling_points(selling_points)}

## 消费者决策因素
{self._fmt_decision_factors(consumer)}

---

# 🎯 任务：卖点选择与排序（主图+6张附图，共7张）

你是亚马逊Listing优化专家。为7张图片确定卖点和顺序。

## 输出要求（中文）

### 主图（第1张）
- **表达卖点**：[最核心]
- **选择理由**：

### 附图1-6（第2-7张）
[每张同上格式]

## 排序逻辑说明

要求：结合消费者决策因素排序，覆盖核心购买因素。"""

        elif stage_num == 1:
            image_plan = GlobalState.get("IMAGE_PLAN") or {}
            return f"""{ps}

## 图片规划
{self._fmt_json(image_plan)}

## 参考产品视觉风格
{self._fmt_reference_visual(reference)}

---

# 🎯 任务：卖点表达方式设计

你是创意设计策略专家。为每张图设计表达方式和AI绘图prompt。

⚠️ **每张图卖点必须严格遵循图片规划，不得更改。**

对每张图输出：

## 主图（第1张）
### 对应卖点
### 表达方式研究
- 同类产品常见表达
- 跨行业案例
- 设计建议
### AI绘图Prompt
```
[详细prompt]
```

[其他6张同上]"""

        else:
            image_plan = GlobalState.get("IMAGE_PLAN") or {}
            image_design = GlobalState.get("IMAGE_DESIGN") or {}
            keywords = GlobalState.get("KEYWORDS") or {}
            top_kw = self._fmt_keywords_top(keywords, 10)
            kw_sec = f"\n**Top 10 关键词（图片文案应融入）：**\n{top_kw}\n" if top_kw else ""

            return f"""{ps}

## 图片规划
{self._fmt_json(image_plan)}

## 表达方式设计
{self._fmt_json(image_design)}

## 消费者洞察
{self._fmt_json(consumer)}
{kw_sec}

---

# 🎯 任务：主图+附图设计Brief

你是设计项目经理。生成完整设计Brief文档。

⚠️ **每张图卖点必须与规划一一对应。**

# 亚马逊Listing图片设计Brief

## 1. 项目概述

## 2. 消费者洞察总结

## 3. 图片设计需求（每张图）
- **核心卖点**
- **图片文案**（英文主标题3-6词+副标题）
- **构图设计**（布局、角度、产品状态、背景）
- **文字排版**
- **色彩方案**
- **AI绘图Prompt**

## 4. 设计规范
## 5. 验收标准"""

    def _build_branch_b_prompt(self, stage_num):
        product_info = GlobalState.get("PRODUCT_INFO") or {}
        selling_points = GlobalState.get("SELLING_POINTS") or {}
        consumer = GlobalState.get("CONSUMER_INSIGHTS") or {}
        usage = GlobalState.get("USAGE_FLOW") or {}
        ps = self._fmt_product_info(product_info)
        brand = product_info.get("brand", "")
        bn = f"\n**⚠️ A+页面必须使用品牌名「{brand}」**\n" if brand else ""

        if stage_num == 0:
            return f"""{ps}
{bn}
## 卖点列表
{self._fmt_selling_points(selling_points)}

## 消费者洞察
{self._fmt_json(consumer)}

## 使用流程
{self._fmt_json(usage)}

---

# 🎯 任务：A+页面模块规划

你是A+页面设计专家。规划5个模块。

⚠️ 所有模块使用**高级A+完整图片模式**（1464×600px 完整设计图）。

每个模块：
### 模块N：[类型]
- **名称**
- **核心目标**
- **表达内容**
- **画面构想**（1464×600px 布局）

## 模块排序逻辑"""

        else:
            modules = GlobalState.get("APLUS_MODULES") or {}
            keywords = GlobalState.get("KEYWORDS") or {}
            top_kw = self._fmt_keywords_top(keywords, 10)
            kw_sec = f"\n**Top 10 关键词（A+文案应融入）：**\n{top_kw}\n" if top_kw else ""

            return f"""{ps}
{bn}
## A+模块规划
{self._fmt_json(modules)}
{kw_sec}

---

# 🎯 任务：A+页面设计Brief

你是A+设计项目经理。

⚠️ 所有模块1464×600px。

# 亚马逊A+页面设计Brief

## 1. 概述

## 2. 模块设计需求
每个模块：
- **格式**：1464×600px
- **核心目标**
- **图片上文案**（英文标题+正文）
- **画面布局**
- **视觉元素**
- **色彩方案**
- **AI绘图Prompt**

## 3. 设计规范
## 4. 验收标准"""

    def _build_branch_c_prompt(self, stage_num):
        product_info = GlobalState.get("PRODUCT_INFO") or {}
        selling_points = GlobalState.get("SELLING_POINTS") or {}
        consumer = GlobalState.get("CONSUMER_INSIGHTS") or {}
        keywords = GlobalState.get("KEYWORDS") or {}
        review = GlobalState.get("REVIEW_EXTRACTION") or {}
        ps = self._fmt_product_info(product_info)

        if stage_num == 0:
            return f"""{ps}

## 关键词数据（全量）
{self._fmt_keywords_full(keywords)}

## 卖点列表
{self._fmt_selling_points(selling_points)}

---

# 🎯 任务：关键词策略分析

你是亚马逊SEO专家。

### 1. 核心主词（标题必含，3-5个）
| 关键词 | 周搜索量 | 理由 |

### 2. 重要长尾词（10-15个）
| 关键词 | 周搜索量 | 建议位置 |

### 3. 五点专用词（按卖点分组）
| 卖点 | 关键词 | 融入方式 |

### 4. 布局建议
标题和每条五点分别含哪些词"""

        elif stage_num == 1:
            kw_strategy = GlobalState.get("KEYWORD_STRATEGY") or {}
            return f"""{ps}

## 关键词策略
{self._fmt_json(kw_strategy)}

## 卖点列表
{self._fmt_selling_points(selling_points)}

## 消费者画像
{self._fmt_decision_factors(consumer)}

---

# 🎯 任务：标题生成

你是亚马逊标题专家。

要求：200字符内、含核心主词、突出1-2卖点、符合亚马逊规范。

### 方案1（推荐）
**标题**：
**字符数**：
**包含关键词**：
**突出卖点**：
**思路**：

### 方案2-3（备选）
[同上]

### 方案对比
| 方案 | 优点 | 缺点 | 推荐度 |"""

        else:
            kw_strategy = GlobalState.get("KEYWORD_STRATEGY") or {}
            title = GlobalState.get("TITLE") or {}
            voc = self._fmt_voc(review)
            voc_sec = f"\n{voc}\n" if voc else ""

            return f"""{ps}

## 关键词策略
{self._fmt_json(kw_strategy)}

## 标题方案
{self._fmt_json(title)}

## 卖点列表
{self._fmt_selling_points(selling_points)}

## 消费者痛点
{self._fmt_pain_points(consumer)}
{voc_sec}

---

# 🎯 任务：五点描述生成

你是亚马逊五点描述专家。

要求：
- 每条500字符内，聚焦一个卖点
- 自然融入关键词
- 结构：大写标题 + 描述
- **尽量使用买家原话（VOC）**
- 差评痛点正面回应

### 五点1-5
每条：**主题** / **关键词** / **内容** / **字符数**

### 关键词覆盖检查
| 关键词 | 位置 |

### Backend Search Terms（≤250字节，不重复已用词）
```
[搜索词]
```
**字节数**："""

    def _get_branch_extraction_config(self, branch, stage_num):
        """返回 (key, prompt, fields)。最终Brief/五点阶段返回 (None, None, None)。"""
        c = {
            ("A", 0): ("IMAGE_PLAN", """提取图片规划。
Output in ```json format:
```json
{"images": [{"position": "主图/附图N", "selling_point": "卖点", "reason": "理由"}], "ordering_logic": "逻辑"}
```""", ["images"]),

            ("A", 1): ("IMAGE_DESIGN", """提取表达方式和AI绘图prompt。
Output in ```json format:
```json
{"designs": [{"position": "主图/附图N", "selling_point": "卖点", "approach": "方式", "ai_prompt": "prompt"}]}
```""", ["designs"]),

            ("A", 2): (None, None, None),

            ("B", 0): ("APLUS_MODULES", """提取A+模块规划。
Output in ```json format:
```json
{"modules": [{"name": "名", "type": "类型", "objective": "目标", "content": "内容", "layout": "布局"}], "ordering_logic": "逻辑"}
```""", ["modules"]),

            ("B", 1): (None, None, None),

            ("C", 0): ("KEYWORD_STRATEGY", """提取关键词策略。
Output in ```json format:
```json
{"core_keywords": [{"keyword": "词", "volume": "量", "reason": "理由"}], "long_tail_keywords": [{"keyword": "词", "volume": "量", "placement": "位置"}], "bullet_keywords": [{"selling_point": "卖点", "keywords": ["词"], "method": "方式"}], "layout_plan": {"title": "", "bullet_1": "", "bullet_2": "", "bullet_3": "", "bullet_4": "", "bullet_5": ""}}
```""", ["core_keywords"]),

            ("C", 1): ("TITLE", """提取标题方案。
Output in ```json format:
```json
{"recommended": {"title": "标题", "char_count": 0, "keywords": ["词"], "selling_points": ["卖点"]}, "alternatives": [{"title": "标题", "char_count": 0}]}
```""", ["recommended"]),

            ("C", 2): (None, None, None),
        }
        return c.get((branch, stage_num), (None, None, None))

    # ============ 分支专用格式化 ============

    def _fmt_selling_points(self, sp):
        if not sp:
            return "[无]"
        if "_raw_text" in sp:
            return sp["_raw_text"]
        points = sp.get("selling_points", [])
        if not points:
            return "[无卖点]"
        lines = []
        for p in points:
            lines.append(f"{p.get('rank','?')}. **{p.get('name','?')}** — {p.get('description','')}")
            if p.get('evidence'):
                lines.append(f"   证据：{p['evidence']}")
            if p.get('vs_competitors'):
                lines.append(f"   竞品对比：{p['vs_competitors']}")
            if p.get('keywords_en'):
                lines.append(f"   英文：{', '.join(p['keywords_en'])}")
        return "\n".join(lines)

    def _fmt_decision_factors(self, consumer):
        if not consumer or "_raw_text" in consumer:
            return self._fmt_json(consumer) if consumer else "[无]"
        factors = consumer.get("decision_factors", [])
        if not factors:
            return "[无]"
        return "\n".join([f"- **{f.get('factor','?')}**（{f.get('importance','?')}）：{f.get('reason','')}" for f in factors])

    def _fmt_pain_points(self, consumer):
        if not consumer or "_raw_text" in consumer:
            return self._fmt_json(consumer) if consumer else "[无]"
        pains = consumer.get("pain_points", [])
        if not pains:
            return "[无]"
        lines = []
        for p in pains:
            line = f"- **{p.get('pain','?')}**（优先级{p.get('priority','?')}）：{p.get('manifestation','')}"
            if p.get('en_quotes'):
                line += f" — {', '.join(p['en_quotes'][:2])}"
            lines.append(line)
        return "\n".join(lines)

    def _fmt_reference_visual(self, ref):
        if not ref or "_raw_text" in ref:
            return self._fmt_json(ref) if ref else "[无]"
        parts = []
        for p in ref.get("products", []):
            if p.get("visual_style"):
                parts.append(f"- {p.get('title','?')[:50]}：{p['visual_style']}")
        return "\n".join(parts) if parts else "[无]"

    def _fmt_keywords_top(self, kw_data, n):
        if not kw_data:
            return ""
        kws = kw_data.get("keywords", [])
        if not kws:
            return ""
        lines = ["| 关键词 | 翻译 | 周搜索量 |", "|--------|------|----------|"]
        for kw in kws[:n]:
            lines.append(f"| {kw.get('keyword','')} | {kw.get('translation','')} | {kw.get('volume',0)} |")
        return "\n".join(lines)

    def _fmt_keywords_full(self, kw_data):
        if not kw_data:
            return "[无关键词]"
        kws = kw_data.get("keywords", [])
        if not kws:
            return "[无]"
        lines = [f"**共 {len(kws)} 个（按搜索量排序）**\n",
                 "| # | 关键词 | 翻译 | 相关性 | 周搜索量 |",
                 "|---|--------|------|--------|----------|"]
        for i, kw in enumerate(kws, 1):
            lines.append(f"| {i} | {kw.get('keyword','')} | {kw.get('translation','')} | {kw.get('relevance','')} | {kw.get('volume',0)} |")
        return "\n".join(lines)

    def _fmt_voc(self, review):
        if not review or "_raw_text" in review:
            return ""
        voc = review.get("voc_gold_quotes", {})
        if not voc:
            return ""
        parts = ["**🗣️ 买家原话（VOC）— 请在五点中融入：**"]
        if voc.get("positive"):
            parts.append(f"- 好评：{', '.join(str(x) for x in voc['positive'][:15])}")
        if voc.get("negative"):
            parts.append(f"- 差评（需正面回应）：{', '.join(str(x) for x in voc['negative'][:10])}")
        if voc.get("scenario"):
            parts.append(f"- 场景词：{', '.join(str(x) for x in voc['scenario'][:10])}")
        return "\n".join(parts)


if __name__ == "__main__":
    bot = ListingDesigner()
    bot.run()